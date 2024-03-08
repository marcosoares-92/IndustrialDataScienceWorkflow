"""FUNCTIONS FROM INDUSTRIAL DATA SCIENCE WORKFLOW (IDSW) PACKAGE
Pipelines for Google Cloud Platform (GCP), Google Colab and Google Drive
Pipelines for Amazon Simple Storage Service (S3)
Pipelines for reading tables and non-structured sheets on Excel
Pipelines for loading Pandas dataframe and exporting as CSV and Excel
Extract data from Plant Information Management (PIMS) systems
AspenTech IP21
Connect to SQLite Database

Marco Cesar Prado Soares, Data Scientist Specialist @ Bayer Crop Science LATAM
marcosoares.feq@gmail.com
marco.soares@bayer.com"""

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
import openpyxl

from dataclasses import dataclass
from idsw import (InvalidInputsError, ControlVars)


@dataclass
class Connectors:
  """
    Store the connectors as Global variables to use, so that the classes do not have to be re-instantiated always.
    Variables are always started with value None or as booleans.

    for a variable var, syntax 
    if var:
        # is equivalent to: "run only when variable var is not None." or "run only if var is True".
    """
  
  # While persistent = True, the system will try to use the same connector already created.
  # User must change this variable state to create new connectors.
  persistent = True
  
  # This class will also store the connectors, once they are created.


class MountGoogleDrive:
    """Class for Mounting Google Drive"""
    def __init__ (self):

        from google.colab import drive
        # Google Colab library must be imported only in case it is
        # going to be used, for avoiding AWS compatibility issues.
        
        print("Associate the Python environment to your Google Drive account, and authorize the access in the opened window.")
        
        drive.mount('/content/drive')
        
        print("Now your Python environment is connected to your Google Drive: the root directory of your environment is now the root of your Google Drive.")
        print("In Google Colab, navigate to the folder icon (\'Files\') of the left navigation menu to find a specific folder or file in your Google Drive.")
        print("Click on the folder or file name and select the elipsis (...) icon on the right of the name to reveal the option \'Copy path\', which will give you the path to use as input for loading objects and files on your Python environment.")
        print("Caution: save your files into different directories of the Google Drive. If files are all saved in a same folder or directory, like the root path, they may not be accessible from your Python environment.")
        print("If you still cannot see the file after moving it to a different folder, reload the environment.")
    

    def upload_to_colab(self):

        from google.colab import files
        # google.colab library must be imported only in case 
        # it is going to be used, for avoiding 
        # AWS compatibility issues.
        print("Click on the button for file selection and select the files from your machine that will be uploaded in the Colab environment.")
        print("Warning: the files will be removed from Colab memory after the Kernel dies or after the notebook is closed.")
        # this functionality requires the previous declaration:
        ## from google.colab import files
            
        colab_files_dict = files.upload()
            
        # The files are stored into a dictionary called colab_files_dict where the keys
        # are the names of the files and the values are the files themselves.
        ## e.g. if you upload a single file named "dictionary.pkl", the dictionary will be
        ## colab_files_dict = {'dictionary.pkl': file}, where file is actually a big string
        ## representing the contents of the file. The length of this value is the size of the
        ## uploaded file, in bytes.
        ## To access the file is like accessing a value from a dictionary: 
        ## d = {'key1': 'val1'}, d['key1'] == 'val1'
        ## we simply declare the key inside brackets and quotes, the same way we would do for
        ## accessing the column of a dataframe.
        ## In this example, colab_files_dict['dictionary.pkl'] access the content of the 
        ## .pkl file, and len(colab_files_dict['dictionary.pkl']) is the size of the .pkl
        ## file in bytes.
        ## To check the dictionary keys, apply the method .keys() to the dictionary (with empty
        ## parentheses): colab_files_dict.keys()
            
        for key in colab_files_dict.keys():
            #loop through each element of the list of keys of the dictionary
            # (list colab_files_dict.keys()). Each element is named 'key'
            print(f"User uploaded file {key} with length {len(colab_files_dict[key])} bytes.")
            # The key is the name of the file, and the length of the value
            ## correspondent to the key is the file's size in bytes.
            ## Notice that the content of the uploaded object must be passed 
            ## as argument for a proper function to be interpreted. 
            ## For instance, the content of a xlsx file should be passed as
            ## argument for Pandas .read_excel function; the pkl file must be passed as
            ## argument for pickle.
            ## e.g., if you uploaded 'table.xlsx' and stored it into colab_files_dict you should
            ## declare df = pd.read_excel(colab_files_dict['table.xlsx']) to obtain a dataframe
            ## df from the uploaded table. Notice that is the value, not the key, that is the
            ## argument.
                
            print("The uploaded files are stored into a dictionary object named as colab_files_dict.")
            print("Each key from this dictionary is the name of an uploaded file. The value correspondent to that key is the file itself.")
            print("The structure of a general Python dictionary is dict = {\'key1\': value1}. To access value1, declare file = dict[\'key1\'], as if you were accessing a column from a dataframe.")
            print("Then, if you uploaded a file named \'table.xlsx\', you can access this file as:")
            print("uploaded_file = colab_files_dict[\'table.xlsx\']")
            print("Notice, though, that the object uploaded_file is the whole file content, not a Python object already converted. To convert to a Python object, pass this element as argument for a proper function or method.")
            print("In this example, to convert the object uploaded_file to a dataframe, Pandas pd.read_excel function could be used. In the following line, a df dataframe object is obtained from the uploaded file:")
            print("df = pd.read_excel(uploaded_file)")
            print("Also, the uploaded file itself will be available in the Colaboratory Notebook\'s workspace.")
            
            self.colab_files_dict =  colab_files_dict

            return self
    

    def download_from_colab(self, file_to_download_from_colab = None):

        from google.colab import files

            
        if (file_to_download_from_colab is None):
                
            #No object was declared
            print("Please, inform a file to download from the notebook\'s workspace. It should be declared in quotes and with the extension: e.g. \'table.csv\'.")
            
        else:
                
            print("The file will be downloaded to your computer.")

            files.download(file_to_download_from_colab)

            print(f"File {file_to_download_from_colab} successfully downloaded from Colab environment.")
        
        return self


class AWSS3Connection:
    """Class for connecting with Amazon AWS Single Storage Service (S3)"""

    
    def __init__ (self, path_to_store_imported_s3_bucket = '', s3_bucket_name = None, s3_obj_prefix = None):

        from getpass import getpass

        self.s3_bucket_name = s3_bucket_name
        self.s3_obj_prefix = s3_obj_prefix

        # Check if path_to_store_imported_s3_bucket is None. If it is, make it the root directory:
        if ((path_to_store_imported_s3_bucket is None)|(str(path_to_store_imported_s3_bucket) == "/")):
            
            # For the S3 buckets, the path should not start with slash. Assign the empty
            # string instead:
            self.path_to_store_imported_s3_bucket = ""
            print("Bucket\'s content will be copied to the notebook\'s root directory.")
        
        elif (str(path_to_store_imported_s3_bucket) == ""):
            # Guarantee that the path is the empty string.
            # Avoid accessing the else condition, what would raise an error
            # since the empty string has no character of index 0
            self.path_to_store_imported_s3_bucket = str(path_to_store_imported_s3_bucket)
            print("Bucket\'s content will be copied to the notebook\'s root directory.")
        
        else:
            # Use the str attribute to guarantee that the path was read as a string:
            path_to_store_imported_s3_bucket = str(path_to_store_imported_s3_bucket)
            
            if(path_to_store_imported_s3_bucket[0] == "/"):
                # the first character is the slash. Let's remove it

                # In AWS, neither the prefix nor the path to which the file will be imported
                # (file from S3 to workspace) or from which the file will be exported to S3
                # (the path in the notebook's workspace) may start with slash, or the operation
                # will not be concluded. Then, we have to remove this character if it is present.

                # The slash is character 0. Then, we want all characters from character 1 (the
                # second) to character len(str(path_to_store_imported_s3_bucket)) - 1, the index
                # of the last character. So, we can slice the string from position 1 to position
                # the slicing syntax is: string[1:] - all string characters from character 1
                # string[:10] - all string characters from character 10-1 = 9 (including 9); or
                # string[1:10] - characters from 1 to 9
                # So, slice the whole string, starting from character 1:
                self.path_to_store_imported_s3_bucket = path_to_store_imported_s3_bucket[1:]
                # attention: even though strings may be seem as list of characters, that can be
                # sliced, we cannot neither simply assign a character to a given position nor delete
                # a character from a position.


    def get_credentials(self):
            
        # Ask the user to provide the credentials:
        ACCESS_KEY = input("Enter your AWS Access Key ID here (in the right). It is the value stored in the field \'Access key ID\' from your AWS user credentials CSV file.")
        print("\n") # line break
        SECRET_KEY = getpass("Enter your password (Secret key) here (in the right). It is the value stored in the field \'Secret access key\' from your AWS user credentials CSV file.")
            
        # The use of 'getpass' instead of 'input' hide the password behind dots.
        # So, the password is not visible by other users and cannot be copied.
            
        print("\n")
        print("WARNING: The bucket\'s name, the prefix, the AWS access key ID, and the AWS Secret access key are all sensitive information, which may grant access to protected information from the organization.\n")
        print("After copying data from S3 to your workspace, remember of removing these information from the notebook, specially if it is going to be shared. Also, remember of removing the files from the workspace.\n")
        print("The cost for storing files in Simple Storage Service is quite inferior than the one for storing directly in SageMaker workspace. Also, files stored in S3 may be accessed for other users than those with access the notebook\'s workspace.\n")

        # Check if the user actually provided the mandatory inputs, instead
        # of putting None or empty string:
        if ((ACCESS_KEY is None) | (ACCESS_KEY == '')):
            raise InvalidInputsError ("AWS Access Key ID is missing. It is the value stored in the field \'Access key ID\' from your AWS user credentials CSV file.")
                
        elif ((SECRET_KEY is None) | (SECRET_KEY == '')):
            raise InvalidInputsError ("AWS Secret Access Key is missing. It is the value stored in the field \'Secret access key\' from your AWS user credentials CSV file.")
                       
        else:
            # Use the str attribute to guarantee that all AWS parameters were properly read as strings, and not as
            # other variables (like integers or floats):
            ACCESS_KEY = str(ACCESS_KEY)
            SECRET_KEY = str(SECRET_KEY)

            # Remove any possible trailing (white and tab spaces) spaces
            # That may be present in the string. Use the Python string
            # rstrip method, which is the equivalent to the Trim function:
            # When no arguments are provided, the whitespaces and tabulations
            # are the removed characters
            # https://www.w3schools.com/python/ref_string_rstrip.asp?msclkid=ee2d05c3c56811ecb1d2189d9f803f65
            self.ACCESS_KEY = ACCESS_KEY.rstrip()
            self.SECRET_KEY = SECRET_KEY.rstrip()
            # Since the user manually inputs the parameters ACCESS and SECRET_KEY,
            # it is easy to input whitespaces without noticing that.
        
        return self


    def get_bucket_info(self):

        if ((self.s3_bucket_name is None) | (self.s3_bucket_name == '')):
            raise InvalidInputsError ("Please, enter a valid S3 Bucket\'s name. Do not add sub-directories or folders (prefixes), only the name of the bucket itself.")
        
        s3_bucket_name = str(self.s3_bucket_name)

        if(s3_bucket_name[0] == "/"):
                # the first character is the slash. Let's remove it

                # In AWS, neither the prefix nor the path to which the file will be imported
                # (file from S3 to workspace) or from which the file will be exported to S3
                # (the path in the notebook's workspace) may start with slash, or the operation
                # will not be concluded. Then, we have to remove this character if it is present.

                # So, slice the whole string, starting from character 1 (as did for 
                # path_to_store_imported_s3_bucket):
                s3_bucket_name = s3_bucket_name[1:]

        self.s3_bucket_name = s3_bucket_name.rstrip()
        
        # Now process the non-obbligatory parameter.
        # Check if a prefix was passed as input parameter. If so, we must select only the names that start with
        # The prefix.
        # Example: in the bucket 'my_bucket' we have a directory 'dir1'.
        # In the main (root) directory, we have a file 'file1.json' like: '/file1.json'
        # If we pass the prefix 'dir1', we want only the files that start as '/dir1/'
        # such as: 'dir1/file2.json', excluding the file in the main (root) directory and excluding the files in other
        # directories. Also, we want to eliminate the file names with no extensions, like 'dir1/' or 'dir1/dir2',
        # since these object names represent folders or directories, not files.

           
        if (self.s3_obj_prefix is None):
            print ("No prefix, specific object, or subdirectory provided.") 
            print (f"Then, exporting to or importing from \'{self.s3_bucket_name}\' root (main) directory.\n")
            # s3_path: path that the file should have in S3:
            self.s3_path = "" # empty string for the root directory
        
        elif ((s3_obj_prefix == "/") | (s3_obj_prefix == '')):
            
            self.s3_obj_prefix = None
            # The root directory in the bucket must not be specified starting with the slash
            # If the root "/" or the empty string '' is provided, make
            # it equivalent to None (no directory)
            print ("No prefix, specific object, or subdirectory provided.") 
            print (f"Then, exporting to or importing from \'{self.s3_bucket_name}\' root (main) directory.\n")
            # s3_path: path that the file should have in S3:
            self.s3_path = "" # empty string for the root directory
        
        else:
            # Since there is a prefix, use the str attribute to guarantee that the path was read as a string:
            s3_obj_prefix = str(self.s3_obj_prefix)
                
            if(s3_obj_prefix[0] == "/"):
                # the first character is the slash. Let's remove it

                # In AWS, neither the prefix nor the path to which the file will be imported
                # (file from S3 to workspace) or from which the file will be exported to S3
                # (the path in the notebook's workspace) may start with slash, or the operation
                # will not be concluded. Then, we have to remove this character if it is present.

                # So, slice the whole string, starting from character 1 (as did for 
                # path_to_store_imported_s3_bucket):
                s3_obj_prefix = s3_obj_prefix[1:]

            # Remove any possible trailing (white and tab spaces) spaces
            # That may be present in the string. Use the Python string
            # rstrip method, which is the equivalent to the Trim function:
            self.s3_obj_prefix = s3_obj_prefix.rstrip()
                
            # s3_path: path that the file should have in S3:
            # Make the path the prefix itself, since there is a prefix:
            self.s3_path = self.s3_obj_prefix
            
            # Store the total characters in the prefix string after removing the initial slash
            # and trailing spaces:
            self.prefix_len = len(self.s3_obj_prefix)
            
            print("AWS Access Credentials, and bucket\'s prefix, object or subdirectory provided.\n")

            return self
    

    def connect_to_s3(self):

        import os
        import boto3
        # boto3 is AWS S3 Python SDK
        # sagemaker and boto3 libraries must be imported only in case 
        # they are going to be used, for avoiding 
        # Google Colab compatibility issues.

        print ("Starting connection with the S3 bucket.\n")
        
        try:
            # Start S3 client as the object 's3_client'
            self.s3_client = boto3.resource('s3', aws_access_key_id = self.ACCESS_KEY, aws_secret_access_key = self.SECRET_KEY)
        
            print(f"Credentials accepted by AWS. S3 client successfully started.\n")
            # An object 'data_table.xlsx' in the main (root) directory of the s3_bucket is stored in Python environment as:
            # s3.ObjectSummary(bucket_name='bucket_name', key='data_table.xlsx')
            # The name of each object is stored as the attribute 'key' of the object.
        
        except:
            
            print("Failed to connect to AWS Simple Storage Service (S3). Review if your credentials are correct.")
            print("The variable \'access_key\' must be set as the value (string) stored as \'Access key ID\' in your user security credentials CSV file.")
            print("The variable \'secret_key\' must be set as the value (string) stored as \'Secret access key\' in your user security credentials CSV file.")
        
        return self
    

    def connect_to_bucket(self):

        try:
            # Connect to the bucket specified as 'bucket_name'.
            # The bucket is started as the object 's3_bucket':
            self.s3_bucket = self.s3_client.Bucket(self.s3_bucket_name)
            print(f"Connection with bucket \'{s3_bucket_name}\' stablished.\n")
            
        except:
            
            print("Failed to connect with the bucket, which usually happens when declaring a wrong bucket\'s name.") 
            print("Check the spelling of your bucket_name string and remember that it must be all in lower-case.\n")
        
        return self
    

    def map_bucket_contents(self):

        import os

         # Then, let's obtain a list of all objects in the bucket (list bucket_objects):
        
        bucket_objects_list = [str(stored_obj.key) for stored_obj in self.s3_bucket.objects.all()]

        # Now get a list to store only the elements from bucket_objects_list that are not folders or directories
        # (objects with extensions).
        # If a prefix was provided, only files with that prefix should be added:

        # Loop through all elements 'stored_obj' from the list bucket_objects_list
        # Check the file extension: file_extension = os.path.splitext(stored_obj)[1][1:]
        self.bucket_objects_list = [stored_obj for stored_obj in bucket_objects_list if (os.path.splitext(stored_obj)[1][1:] != '')]

        # The os.path.splitext method splits the string into its FIRST dot (".") to
        # separate the file extension from the full path. Example:
        # "C:/dir1/dir2/data_table.csv" is split into:
        # "C:/dir1/dir2/data_table" (root part) and '.csv' (extension part)
        # https://www.geeksforgeeks.org/python-os-path-splitext-method/?msclkid=2d56198fc5d311ec820530cfa4c6d574

        # os.path.splitext(stored_obj) is a tuple of strings: the first is the complete file
        # root with no extension; the second is the extension starting with a point: '.txt'
        # When we set os.path.splitext(stored_obj)[1], we are selecting the second element of
        # the tuple. By selecting os.path.splitext(stored_obj)[1][1:], we are taking this string
        # from the second character (index 1), eliminating the dot: 'txt'

        if not (self.s3_obj_prefix is None):
                        
            # Check the characters from the position 0 (1st character) to the position
            # prefix_len - 1. Since a prefix was declared, we want only the objects that this first portion
            # corresponds to the prefix. string[i:j] slices the string from index i to index j-1
            # Then, the 1st portion of the string to check is: string[0:(prefix_len)]

            # Slice the string stored_obj from position 0 (1st character) to position prefix_len - 1,
            # The position that the prefix should end: obj_name_first_part = (stored_obj)[0:(self.prefix_len)]
            # If this first part is the prefix, then append the object to list:
            self.bucket_objects_list = [stored_obj for stored_obj in bucket_objects_list if ((stored_obj)[0:(self.prefix_len)] == (self.s3_obj_prefix))]
           
        # Now, bucket_objects_list contains the names of all objects from the bucket that must be copied.

        print("Finished mapping objects to fetch. Now, all these objects from S3 bucket will be copied to the notebook\'s workspace, in the specified directory.\n")
        print(f"A total of {len(self.bucket_objects_list)} files were found in the specified bucket\'s prefix (\'{self.s3_obj_prefix}\').")
        print(f"The first file found is \'{self.bucket_objects_list[0]}\'; whereas the last file found is \'{self.bucket_objects_list[len(self.bucket_objects_list) - 1]}\'.")
            
        return self


    def copy_bucket_files(self):

        import os
        # Now, let's try copying the files:
            
        try:
            
            self.objects_to_copy = [self.s3_bucket.Object(copied_object) for copied_object in self.bucket_objects_list]

            # Now, copy objects to the workspace:
            # Set the new file_path. Notice that by now, copied_object may be a string like:
            # 'dir1/.../dirN/file_name.ext', where dirN is the n-th directory and ext is the file extension.
            # We want only the file_name to joing with the path to store the imported bucket. So, we can use the
            # str.split method specifying the separator sep = '/' to break the string into a list of substrings.
            # The last element from this list will be 'file_name.ext'
            # https://www.w3schools.com/python/ref_string_split.asp?msclkid=135399b6c63111ecada75d7d91add056

            # 1. Break the copied_object full path into the list object_path_list, using the .split method:
            objects_paths_lists = [copied_object.split(sep = "/") for copied_object in self.bucket_objects_list]
            # each element in objects_paths_lists is a list of splitted strings containing information on the paths.
            # 2. Get the last element from this list. Since it has length len(object_path_list) and indexing starts from
            # zero, the index of the last element is (len(object_path_list) - 1):
            self.fetched_objects = [object_path_list[(len(object_path_list) - 1)] for object_path_list in objects_paths_lists]
            # 3. Finally, join the string fetched_object with the new path (path on the notebook's workspace) to finish
            # The new object's file_path:
            self.file_paths = [os.path.join(self.path_to_store_imported_s3_bucket, fetched_object) for fetched_object in fetched_objects]

            # Now, loop through the correspondents objects and paths to download them:
            for selected_object, fetched_object, file_path in zip(self.objects_to_copy, self.fetched_objects, self.file_paths):
                # Download the selected object to the workspace in the specified file_path
                # The parameter Filename must be input with the path of the copied file, including its name and
                # extension. Example Filename = "/my_table.xlsx" copies a xlsx file named 'my_table' to the notebook's main (root)
                # directory
                selected_object.download_file(Filename = file_path)
                print(f"The file \'{fetched_object}\' was successfully copied to notebook\'s workspace.\n")

            print("Finished copying the files from the bucket to the notebook\'s workspace. It may take a couple of minutes untill they be shown in SageMaker environment.\n") 
            print("Do not forget to delete these copies after finishing the analysis. They will remain stored in the bucket.\n")


        except:

            # Run this code for any other exception that may happen (no exception error
            # specified, so any exception runs the following code).
            # Check: https://pythonbasics.org/try-except/?msclkid=4f6b4540c5d011ecb1fe8a4566f632a6
            # for seeing how to handle successive exceptions

            print("Attention! The function raised an exception error, which is probably due to the AWS Simple Storage Service (S3) permissions.")
            print("Before running again this function, check this quick guide for configuring the permission roles in AWS.\n")
            print("It is necessary to create an user with full access permissions to interact with S3 from SageMaker. To configure the User, go to the upper ribbon of AWS, click on Services, and select IAM – Identity and Access Management.")
            print("1. In IAM\'s lateral panel, search for \'Users\' in the group of Access Management.")
            print("2. Click on the \'Add users\' button.")
            print("3. Set an user name in the text box \'User name\'.")
            print("Attention: users and S3 buckets cannot be written in upper case. Also, selecting a name already used by an Amazon user or bucket will raise an error message.\n")
            print("4. In the field \'Select type of Access to AWS\'-\'Select type of AWS credentials\' select the option \'Access key - Programmatic access\'. After that, click on the button \'Next: Permissions\'.")
            print("5. In the field \'Set Permissions\', keep the \'Add user to a group\' button marked.")
            print("6. In the field \'Add user to a group\', click on \'Create group\' (alternatively, you can be added to a group already configured or copy the permissions of another user.")
            print("7. In the text box \'Group\'s name\', set a name for the new group of permissions.")
            print("8. In the search bar below (\'Filter politics\'), search for a politics that fill your needs, and check the option button on the left of this politic. The politics \'AmazonS3FullAccess\' grants full access to the S3 content.")
            print("9. Finally, click on \'Create a group\'.")
            print("10. After the group is created, it will appear with a check box marked, over the previous groups. Keep it marked and click on the button \'Next: Tags\'.")
            print("11. Create and note down the Access key ID and Secret access key. You can also download a comma separated values (CSV) file containing the credentials for future use.")
            print("ATTENTION: These parameters are required for accessing the bucket\'s content from any application, including AWS SageMaker.")
            print("12. Click on \'Next: Review\' and review the user credentials information and permissions.")
            print("13. Click on \'Create user\' and click on the download button to download the CSV file containing the user credentials information.")
            print("The headers of the CSV file (the stored fields) is: \'User name, Password, Access key ID, Secret access key, Console login link\'.")
            print("You need both the values indicated as \'Access key ID\' and as \'Secret access key\' to fetch the S3 bucket.")
            print("\n") # line break
            print("After acquiring the necessary user privileges, use the boto3 library to fetch the bucket from the Python code. boto3 is AWS S3 Python SDK.")
            print("For fetching a specific bucket\'s file use the following code:\n")
            print("1. Set a variable \'access_key\' as the value (string) stored as \'Access key ID\' in your user security credentials CSV file.")
            print("2. Set a variable \'secret_key\' as the value (string) stored as \'Secret access key\' in your user security credentials CSV file.")
            print("3. Set a variable \'bucket_name\' as a string containing only the name of the bucket. Do not add subdirectories, folders (prefixes), or file names.")
            print("Example: if your bucket is named \'my_bucket\' and its main directory contains folders like \'folder1\', \'folder2\', etc, do not declare bucket_name = \'my_bucket/folder1\', even if you only want files from folder1.")
            print("ALWAYS declare only the bucket\'s name: bucket_name = \'my_bucket\'.")
            print("4. Set a variable \'file_path\' containing the path from the bucket\'s subdirectories to the file you want to fetch. Include the file name and its extension.")
            print("If the file is stored in the bucket\'s root (main) directory: file_path = \"my_file.ext\".")
            print("If the path of the file in the bucket is: \'dir1/…/dirN/my_file.ext\', where dirN is the N-th subdirectory, and dir1 is a folder or directory of the main (root) bucket\'s directory: file_path = \"dir1/…/dirN/my_file.ext\".")
            print("Also, we say that \'dir1/…/dirN/\' is the file\'s prefix. Notice that the name of the bucket is never declared here as the path for fetching its content from the Python code.")
            print("5. Set a variable named \'new_path\' to store the path of the file copied to the notebook’s workspace. This path must contain the file name and its extension.")
            print("Example: if you want to copy \'my_file.ext\' to the root directory of the notebook’s workspace, set: new_path = \"/my_file.ext\".")
            print("6. Finally, declare the following code, which refers to the defined variables:\n")

            # Let's use triple quotes to declare a formated string
            example_code = """
                import boto3
                # Start S3 client as the object 's3_client'
                s3_client = boto3.resource('s3', aws_access_key_id = access_key, aws_secret_access_key = secret_key)
                # Connect to the bucket specified as 'bucket_name'.
                # The bucket is started as the object 's3_bucket':
                s3_bucket = s3_client.Bucket(bucket_name)
                # Select the object in the bucket previously started as 's3_bucket':
                selected_object = s3_bucket.Object(file_path)
                # Download the selected object to the workspace in the specified file_path
                # The parameter Filename must be input with the path of the copied file, including its name and
                # extension. Example Filename = "/my_table.xlsx" copies a xlsx file named 'my_table' to the notebook's main (root)
                # directory
                selected_object.download_file(Filename = new_path)
                """

            print(example_code)

            print("An object \'my_file.ext\' in the main (root) directory of the s3_bucket is stored in Python environment as:")
            print("""s3.ObjectSummary(bucket_name='bucket_name', key='my_file.ext'""") 
            # triple quotes to keep the internal quotes without using too much backslashes "\" (the ignore next character)
            print("Then, the name of each object is stored as the attribute \'key\' of the object. To view all objects, we can loop through their \'key\' attributes:\n")
            example_code = """
                # Loop through all objects of the bucket:
                for stored_obj in s3_bucket.objects.all():		
                    # Loop through all elements 'stored_obj' from s3_bucket.objects.all()
                    # Which stores the ObjectSummary for all objects in the bucket s3_bucket:
                    # Print the object’s names:
                    print(stored_obj.key)
                    """

            print(example_code)
        
        return self
        
    
    def run_s3_connection_pipeline (self):

        self = self.get_credentials()
        self = self.get_bucket_info()
        self = self.connect_to_s3()
        self = self.connect_to_bucket()

        return self
    

    def fetch_s3_files_pipeline (self):

        self = self.map_bucket_contents()
        self = self.copy_bucket_files()

        return self
    

    def set_directory_to_export(self, directory_of_notebook_workspace_storing_files_to_export = None):

        # Check if directory_of_notebook_workspace_storing_files_to_export is None. 
        # If it is, make it the root directory:
        if ((directory_of_notebook_workspace_storing_files_to_export is None)|(str(directory_of_notebook_workspace_storing_files_to_export) == "/")):
                
                # For the S3 buckets, the path should not start with slash. Assign the empty
                # string instead:
                self.directory_of_notebook_workspace_storing_files_to_export = ""
                print("The files will be exported from the notebook\'s root directory to S3.")
        
        elif (str(directory_of_notebook_workspace_storing_files_to_export) == ""):
            
                # Guarantee that the path is the empty string.
                # Avoid accessing the else condition, what would raise an error
                # since the empty string has no character of index 0
                self.directory_of_notebook_workspace_storing_files_to_export = str(directory_of_notebook_workspace_storing_files_to_export)
                print("The files will be exported from the notebook\'s root directory to S3.")
            
        else:
            # Use the str attribute to guarantee that the path was read as a string:
            self.directory_of_notebook_workspace_storing_files_to_export = str(directory_of_notebook_workspace_storing_files_to_export)
                
            if(directory_of_notebook_workspace_storing_files_to_export[0] == "/"):
                # the first character is the slash. Let's remove it

                # In AWS, neither the prefix nor the path to which the file will be imported
                # (file from S3 to workspace) or from which the file will be exported to S3
                # (the path in the notebook's workspace) may start with slash, or the operation
                # will not be concluded. Then, we have to remove this character if it is present.

                # The slash is character 0. Then, we want all characters from character 1 (the
                # second) to character len(str(path_to_store_imported_s3_bucket)) - 1, the index
                # of the last character. So, we can slice the string from position 1 to position
                # the slicing syntax is: string[1:] - all string characters from character 1
                # string[:10] - all string characters from character 10-1 = 9 (including 9); or
                # string[1:10] - characters from 1 to 9
                # So, slice the whole string, starting from character 1:
                self.directory_of_notebook_workspace_storing_files_to_export = self.directory_of_notebook_workspace_storing_files_to_export[1:]
                # attention: even though strings may be seem as list of characters, that can be
                # sliced, we cannot neither simply assign a character to a given position nor delete
                # a character from a position.
        
        return self
    

    def set_files_to_export(self, list_of_file_names_with_extensions):
        
        import os
        # Now, let's obtain the lists of all file paths in the notebook's workspace and
        # of the paths that the files should have in S3, after being exported.

        self.list_of_file_names_with_extensions = list_of_file_names_with_extensions

        try:
            # Get the full path in the notebook's workspace:
            self.workspace_full_paths = [os.path.join(self.directory_of_notebook_workspace_storing_files_to_export, my_file) for my_file in list_of_file_names_with_extensions]
            # Get the full path that the file will have in S3:
            self.s3_full_paths = [os.path.join(self.s3_path, my_file) for my_file in list_of_file_names_with_extensions]
            # Now, both lists have the same number of elements. For an element (file) i,
            # workspace_full_paths has the full file path in notebook's workspace, and
            # s3_full_paths has the path that the new file should have in S3 bucket.

        except:
            raise InvalidInputsError ("The function returned an error when trying to access the list of files. Declare it as a list of strings, even if there is a single element in the list. Example: list_of_file_names_with_extensions = [\'my_file.ext\']\n")
        
        return self
    

    def export_files(self):
        
        # Now, loop through all elements i from the lists.
        # The first elements of the lists have index 0; the last elements have index
        # total_of_files - 1, since there are 'total_of_files' elements:
        
        # Then, export the correspondent element to S3:
        
        try:

            s3_objects = [self.s3_bucket.Object(S3_FILE_PATH) for S3_FILE_PATH in s3_full_paths]

            # Now, loop through the correspondents objects and paths to download them:
            for new_s3_object, uploaded_object, PATH_IN_WORKSPACE in zip(self.s3_objects, self.list_of_file_names_with_extensions, self.workspace_full_paths):
                # Upload the selected object from the workspace path PATH_IN_WORKSPACE
                # to the S3 path specified as S3_FILE_PATH.
                # The parameter Filename must be input with the path of the copied file, including its name and
                # extension. Example Filename = "/my_table.xlsx" exports a xlsx file named 'my_table' to the notebook's main (root)
                # directory
                new_s3_object.upload_file(Filename = PATH_IN_WORKSPACE)
                print(f"The file \'{uploaded_object}\' was successfully exported from notebook\'s workspace to AWS Simple Storage Service (S3).\n")
                
            print("Finished exporting the files from the the notebook\'s workspace to S3 bucket. It may take a couple of minutes untill they be shown in S3 environment.\n") 
            print("Do not forget to delete these copies after finishing the analysis. They will remain stored in the bucket.\n")


        except:

            # Run this code for any other exception that may happen (no exception error
            # specified, so any exception runs the following code).
            # Check: https://pythonbasics.org/try-except/?msclkid=4f6b4540c5d011ecb1fe8a4566f632a6
            # for seeing how to handle successive exceptions

            print("Attention! The function raised an exception error, which is probably due to the AWS Simple Storage Service (S3) permissions.")
            print("Before running again this function, check this quick guide for configuring the permission roles in AWS.\n")
            print("It is necessary to create an user with full access permissions to interact with S3 from SageMaker. To configure the User, go to the upper ribbon of AWS, click on Services, and select IAM – Identity and Access Management.")
            print("1. In IAM\'s lateral panel, search for \'Users\' in the group of Access Management.")
            print("2. Click on the \'Add users\' button.")
            print("3. Set an user name in the text box \'User name\'.")
            print("Attention: users and S3 buckets cannot be written in upper case. Also, selecting a name already used by an Amazon user or bucket will raise an error message.\n")
            print("4. In the field \'Select type of Access to AWS\'-\'Select type of AWS credentials\' select the option \'Access key - Programmatic access\'. After that, click on the button \'Next: Permissions\'.")
            print("5. In the field \'Set Permissions\', keep the \'Add user to a group\' button marked.")
            print("6. In the field \'Add user to a group\', click on \'Create group\' (alternatively, you can be added to a group already configured or copy the permissions of another user.")
            print("7. In the text box \'Group\'s name\', set a name for the new group of permissions.")
            print("8. In the search bar below (\'Filter politics\'), search for a politics that fill your needs, and check the option button on the left of this politic. The politics \'AmazonS3FullAccess\' grants full access to the S3 content.")
            print("9. Finally, click on \'Create a group\'.")
            print("10. After the group is created, it will appear with a check box marked, over the previous groups. Keep it marked and click on the button \'Next: Tags\'.")
            print("11. Create and note down the Access key ID and Secret access key. You can also download a comma separated values (CSV) file containing the credentials for future use.")
            print("ATTENTION: These parameters are required for accessing the bucket\'s content from any application, including AWS SageMaker.")
            print("12. Click on \'Next: Review\' and review the user credentials information and permissions.")
            print("13. Click on \'Create user\' and click on the download button to download the CSV file containing the user credentials information.")
            print("The headers of the CSV file (the stored fields) is: \'User name, Password, Access key ID, Secret access key, Console login link\'.")
            print("You need both the values indicated as \'Access key ID\' and as \'Secret access key\' to fetch the S3 bucket.")
            print("\n") # line break
            print("After acquiring the necessary user privileges, use the boto3 library to export the file from the notebook’s workspace to the bucket (i.e., to upload a file to the bucket).")
            print("For exporting the file as a new bucket\'s file use the following code:\n")
            print("1. Set a variable \'access_key\' as the value (string) stored as \'Access key ID\' in your user security credentials CSV file.")
            print("2. Set a variable \'secret_key\' as the value (string) stored as \'Secret access key\' in your user security credentials CSV file.")
            print("3. Set a variable \'bucket_name\' as a string containing only the name of the bucket. Do not add subdirectories, folders (prefixes), or file names.")
            print("Example: if your bucket is named \'my_bucket\' and its main directory contains folders like \'folder1\', \'folder2\', etc, do not declare bucket_name = \'my_bucket/folder1\', even if you only want files from folder1.")
            print("ALWAYS declare only the bucket\'s name: bucket_name = \'my_bucket\'.")
            print("4. Set a variable \'file_path_in_workspace\' containing the path of the file in notebook’s workspace. The file will be exported from “file_path_in_workspace” to the S3 bucket.")
            print("If the file is stored in the notebook\'s root (main) directory: file_path = \"my_file.ext\".")
            print("If the path of the file in the notebook workspace is: \'dir1/…/dirN/my_file.ext\', where dirN is the N-th subdirectory, and dir1 is a folder or directory of the main (root) bucket\'s directory: file_path = \"dir1/…/dirN/my_file.ext\".")
            print("5. Set a variable named \'file_path_in_s3\' containing the path from the bucket’s subdirectories to the file you want to fetch. Include the file name and its extension.")
            print("6. Finally, declare the following code, which refers to the defined variables:\n")

            # Let's use triple quotes to declare a formated string
            example_code = """
                import boto3
                # Start S3 client as the object 's3_client'
                s3_client = boto3.resource('s3', aws_access_key_id = access_key, aws_secret_access_key = secret_key)
                # Connect to the bucket specified as 'bucket_name'.
                # The bucket is started as the object 's3_bucket':
                s3_bucket = s3_client.Bucket(bucket_name)
                # Start the new object in the bucket previously started as 's3_bucket'.
                # Start it with the specified prefix, in file_path_in_s3:
                new_s3_object = s3_bucket.Object(file_path_in_s3)
                # Finally, upload the file in file_path_in_workspace.
                # Make new_s3_object the exported file:
                # Upload the selected object from the workspace path file_path_in_workspace
                # to the S3 path specified as file_path_in_s3.
                # The parameter Filename must be input with the path of the copied file, including its name and
                # extension. Example Filename = "/my_table.xlsx" exports a xlsx file named 'my_table' to 
                # the notebook's main (root) directory.
                new_s3_object.upload_file(Filename = file_path_in_workspace)
                """

            print(example_code)

            print("An object \'my_file.ext\' in the main (root) directory of the s3_bucket is stored in Python environment as:")
            print("""s3.ObjectSummary(bucket_name='bucket_name', key='my_file.ext'""") 
            # triple quotes to keep the internal quotes without using too much backslashes "\" (the ignore next character)
            print("Then, the name of each object is stored as the attribute \'key\' of the object. To view all objects, we can loop through their \'key\' attributes:\n")
            example_code = """
                # Loop through all objects of the bucket:
                for stored_obj in s3_bucket.objects.all():		
                    # Loop through all elements 'stored_obj' from s3_bucket.objects.all()
                    # Which stores the ObjectSummary for all objects in the bucket s3_bucket:
                    # Print the object’s names:
                    print(stored_obj.key)
                    """

            print(example_code)


        return self


    def export_to_s3_pipeline(self, list_of_file_names_with_extensions, directory_of_notebook_workspace_storing_files_to_export = None):

        self = self.set_directory_to_export(directory_of_notebook_workspace_storing_files_to_export)
        self = self.set_files_to_export(list_of_file_names_with_extensions)
        self = self.export_files()

        return self


class IP21Extractor:
    """
    Class for extracting information from Aspentech IP21 database.
    def __init__ (self, tag_to_extract = None, actual_tag_name = None, ip21_server = None, 
                data_source = 'localhost', start_timestamp = None, stop_timestamp = None, ip21time_array = [], 
                previous_df_for_concatenation = None, username = None, password = None):  
    """    

    # Initialize instance attributes.
    # define the Class constructor, i.e., how are its objects:

    def __init__ (self, previous_df_for_concatenation):
        
        # If the user passes the argument, use them. Otherwise, use the standard values.
        # Set the class objects' attributes.
        # Suppose the object is named assistant. We can access the attribute as:
        # assistant.assistant_startup, for instance.
        # So, we can save the variables as objects' attributes.
        
        # Check if there is a previous dataset for concatenating with new data:
        self.dataset = previous_df_for_concatenation
                
    # Define the class methods.
    # All methods must take an object from the class (self) as one of the parameters


    def get_credentials (self, server, data_source, username, password):
     
        from getpass import getpass
        
        if (username is None):
            
            username = input("Enter your username: ")
            
        if (password is None):
            
            password = getpass("Enter your password: ")
        
        # Remove trailing whitespaces:
        username = str(username).strip()
        password = str(password).strip()
        
        if (data_source is None):
            data_source = 'localhost'
        
        # Attention: do not include http:// in the server, only the server name
        # (what appears after http://)
        self.server = ip21_server
        
        # If no specific data source is provided, use 'localhost'
        self.data_source = data_source
        
        # Create an attribute that checks if another API call is needed:
        self.need_next_call = True
       

        return self
    

    def set_query_parameters (self, tag_to_extract = None, actual_tag_name = None, start_timestamp = None, stop_timestamp = None, ip21time_array = [], previous_df_for_concatenation = None):

        self.tag = tag_to_extract
        # If actual_tag_name is None, make it equal to the tag:
        if (actual_tag_name is None):
            actual_tag_name = tag_to_extract
        
        self.actual_tag_name = actual_tag_name
        
        self.start_timestamp = start_timestamp
        self.start_ip21_scale = np.nan # float, instead of None object
        self.stop_timestamp = stop_timestamp
        self.stop_ip21_scale = np.nan # float, instead of None object
        
        if (ip21time_array is None):
            ip21time_array = []
        
        self.ip21time_array = np.array(ip21time_array)

        return self
    

    def convert_window_to_ip21_timescale (self):
        
        start_timestamp = self.start_timestamp 
        stop_timestamp = self.stop_timestamp
        
        # Pick the closest reference timestamp:
        if (start_timestamp >= pd.Timestamp('06-21-2022 0:00:00.001', unit = 'ns')):
            
            reference = pd.Timestamp('06-21-2022 0:00:00.001', unit = 'ns')
            # reference timestamp in IP21 scale:
            reference_ip21 = 1655780400001
        
        elif (start_timestamp >= pd.Timestamp('06-21-2022', unit = 'd')):
            
            reference = pd.Timestamp('06-21-2022', unit = 'd')
            # reference timestamp in IP21 scale:
            reference_ip21 = 1655780400000
        
        elif (start_timestamp >= pd.Timestamp('06-20-2022', unit = 'd')):
            
            reference = pd.Timestamp('06-20-2022', unit = 'd')
            # reference timestamp in IP21 scale:
            reference_ip21 = 1655694000000
        
        elif (start_timestamp >= pd.Timestamp('01-01-2018', unit = 'd')):
            
            reference = pd.Timestamp('01-01-2018', unit = 'd')
            # reference timestamp in IP21 scale:
            reference_ip21 = 1514772000000
        
        elif (start_timestamp >= pd.Timestamp('01-01-2000', unit = 'd')):
            
            reference = pd.Timestamp('01-01-2000', unit = 'd')
            # reference timestamp in IP21 scale:
            reference_ip21 = 946692000000
        
        elif (start_timestamp >= pd.Timestamp('01-01-1970', unit = 'd')):
            
            reference = pd.Timestamp('01-01-1970', unit = 'd')
            # reference timestamp in IP21 scale:
            reference_ip21 = 10800000
        
        else:
            # Use the lowest timestamp:
            reference = pd.Timestamp('01-01-1960', unit = 'd')
            # reference timestamp in IP21 scale:
            reference_ip21 = -315608400000
            
        
        # Convert the start timestamp:
        start_timedelta = start_timestamp - reference
        # apply the delta method to convert to nanoseconds:
        # The .delta attribute was replaced by .value attribute. 
        # Both return the number of nanoseconds as an integer.
        # https://pandas.pydata.org/docs/reference/api/pandas.Timedelta.html
        start_timedelta = start_timedelta.value
        # 1ms = 10^-3 s, 1ns = 10^-9 s, so 1 ns = 1ms/(10^6)
        # Divide by 10^6 to obtain the total of milliseconds:
        start_timedelta = start_timedelta/(10**6)
        # Sum with the reference value in IP21 scale to obtain the converted timestamp:
        start_ip21_scale = reference_ip21 + start_timedelta
        # Guarantee that the number is an integer:
        # np.rint rounds to the nearest integer, whereas int to convert to integer:
        start_ip21_scale = int(np.rint(start_ip21_scale))
        
        # Convert the stop timestamp:
        stop_timedelta = stop_timestamp - reference
        # apply the delta method to convert to nanoseconds:
        stop_timedelta = stop_timedelta.value
        # Divide by 10^6 to obtain the total of milliseconds:
        stop_timedelta = stop_timedelta/(10**6)
        # Sum with the reference value in IP21 scale to obtain the converted timestamp:
        stop_ip21_scale = reference_ip21 + stop_timedelta
        # Guarantee that the number is an integer:
        # np.rint rounds to the nearest integer, whereas int to convert to integer:
        stop_ip21_scale = int(np.rint(stop_ip21_scale))
        
        # Update the attributes:
        self.start_ip21_scale = start_ip21_scale
        self.stop_ip21_scale = stop_ip21_scale
        
        return self
            
        
    def convert_ip21_timescale_array_to_timestamp (self):
        
        # Convert to Pandas series:
        ip21time_window = pd.Series(self.ip21time_array)
        # Guarantee that the series is sorted ascendingly:
        #https://pandas.pydata.org/pandas-docs/stable/reference/api/pandas.Series.sort_values.html
        ip21time_window = ip21time_window.sort_values(ascending = True)
        # Get the first ip21 time from the series:
        start_ip21_scale = ip21time_window[0]
        
        # Pick the closest reference timestamp:
        if (start_ip21_scale >= 1655780400001):
            
            reference = pd.Timestamp('06-21-2022 0:00:00.001', unit = 'ns')
            # reference timestamp in IP21 scale:
            reference_ip21 = 1655780400001
        
        elif (start_ip21_scale >= 1655780400000):
            
            reference = pd.Timestamp('06-21-2022', unit = 'd')
            # reference timestamp in IP21 scale:
            reference_ip21 = 1655780400000
        
        elif (start_ip21_scale >= 1655694000000):
            
            reference = pd.Timestamp('06-20-2022', unit = 'd')
            # reference timestamp in IP21 scale:
            reference_ip21 = 1655694000000
        
        elif (start_ip21_scale >= 1514772000000):
            
            reference = pd.Timestamp('01-01-2018', unit = 'd')
            # reference timestamp in IP21 scale:
            reference_ip21 = 1514772000000
        
        elif (start_ip21_scale >= 946692000000):
            
            reference = pd.Timestamp('01-01-2000', unit = 'd')
            # reference timestamp in IP21 scale:
            reference_ip21 = 946692000000
        
        elif (start_ip21_scale >= 10800000):
            
            reference = pd.Timestamp('01-01-1970', unit = 'd')
            # reference timestamp in IP21 scale:
            reference_ip21 = 10800000
        
        else:
            # Use the lowest timestamp:
            reference = pd.Timestamp('01-01-1960', unit = 'd')
            # reference timestamp in IP21 scale:
            reference_ip21 = -315608400000
        
        # Get the IP21 timedelta series:
        ip21time_timedeltas = ip21time_window - reference_ip21
        
        # Start a list for the new timestamps:
        new_timestamps = []
        
        # Now, loop through each element from the series ip21time_timedeltas:
        for ip21time_timedelta in ip21time_timedeltas:
            
            # Create a pandas timedelta object, in ms:
            timedelta_obj = pd.Timedelta(ip21time_timedelta, 'ms')
            # Sum this timedelta to reference to obtain the new timestamp:
            new_timestamp = reference + timedelta_obj
            # Append to the list of new_timestamps:
            new_timestamps.append(new_timestamp)
        
        # Now, convert the list to Pandas series:
        timestamp_series = pd.Series(new_timestamps)
        # Rename this series:
        # https://pandas.pydata.org/pandas-docs/stable/reference/api/pandas.Series.rename.html
        timestamp_series = timestamp_series.rename("timestamps")
        
        # Save it as an attribute and return the object:
        self.timestamp_series = timestamp_series
        
        return self
        
    
    def set_extracted_time_window (self, start_timedelta_unit = 'day', stop_timedelta_unit = 'day'):
    
        from datetime import datetime, timedelta

        # start_time: dictionary containing start timestamp information.
        # stop_time: dictionary containing stop timestamp information.

        # Alternatively: start_time = 'today', 'now', start_time = 'yesterday', start_time = -10 for 10
        # days before, start_time = -X for - X days before. Units for offsets will be always in days, unless
        # you modify the parameters start_timedelta_unit and stop_timedelta_unit.
        # For the timedelta unit, set 'day' or 'd' for subtracting values in days,'hour' or 'h',
        # 'minute' or 'm' for minutes, 'second' or 's' for seconds, 'millisecond' or 'ms' for milliseconds.
        # Put the "-" signal, or the time will be interpreted as a future day from today.
        # Analogously for stop_time.
        # Both dictionaries must contain only float values (for 'year', 'day' and 'month'
        # are integers, naturally).

        ## WARNING: The keys must be always be the same, only change the numeric values.
        ## The keys must be: 'year', 'month', 'day', 'hour', 'minute', and 'second'
        
        start_time = self.start_timestamp
        stop_time = self.stop_timestamp
                
        now = datetime.now()
        today = pd.Timestamp(now, unit = 'ns')

        
        if (type (start_time) == dict):

            # Retrieve information from the dictionary:  
            # START timestamp information:
            START_YEAR = start_time['year']
            START_MONTH = start_time['month']
            START_DAY = start_time['day']
            START_HOUR = start_time['hour']
            START_MINUTE = start_time['minute']
            START_SECOND = start_time['second']

            # Create datetime objects from the input information 
            # START_DATETIME: this is the first moment of the database that will be queried.
            START_DATETIME = datetime(year = START_YEAR, month = START_MONTH, day = START_DAY, hour = START_HOUR, minute = START_MINUTE, second = START_SECOND)

            # convert the datetimes to Pandas timestamps, more powerful and compatible with other 
            # Pandas functions, classes and methods. Specify unit = 'ns' (nanoseconds) to guarantee 
            # the resolution
            START_TIMESTAMP = pd.Timestamp(START_DATETIME, unit = 'ns')

        elif (type (start_time) == str):

            if ((start_time == "today") | (start_time == "now")):

                START_TIMESTAMP = today

            elif (start_time == "yesterday"):

                delta_t = pd.Timedelta(1, 'd')
                START_TIMESTAMP = today - delta_t

        elif  ((type (start_time) == float) | (type (start_time) == int)):

                if ((start_timedelta_unit == 'day') | (start_timedelta_unit == 'd')):

                    UNIT = 'd'

                elif ((start_timedelta_unit == 'hour') | (start_timedelta_unit == 'h')):

                    UNIT = 'h'

                elif ((start_timedelta_unit == 'minute') | (start_timedelta_unit == 'm')):

                    UNIT = 'm'

                elif ((start_timedelta_unit == 'second') | (start_timedelta_unit == 's')):

                    UNIT = 's'

                elif ((start_timedelta_unit == 'millisecond') | (start_timedelta_unit == 'ms')):

                    UNIT = 'ms'

                delta_t = pd.Timedelta(start_time, UNIT)
                START_TIMESTAMP = today - delta_t

        if (type (stop_time) == dict):

            # STOP timestamp information:
            STOP_YEAR = stop_time['year']
            STOP_MONTH = stop_time['month']
            STOP_DAY = stop_time['day']
            STOP_HOUR = stop_time['hour']
            STOP_MINUTE = stop_time['minute']
            STOP_SECOND = stop_time['second']

            # STOP_DATETIME: this is the last moment of the database that will be queried.
            STOP_DATETIME = datetime(year = STOP_YEAR, month = STOP_MONTH, day = STOP_DAY, hour = STOP_HOUR, minute = STOP_MINUTE, second = STOP_SECOND)
            STOP_TIMESTAMP = pd.Timestamp(STOP_DATETIME, unit = 'ns')
        
        elif (type (stop_time) == str):

            now = datetime.now()
            today = pd.Timestamp(now, unit = 'ns')

            if ((stop_time == "today") | (stop_time == "now")):

                STOP_TIMESTAMP = today

            elif (stop_time == "yesterday"):

                delta_t = pd.Timedelta(1, 'd')
                STOP_TIMESTAMP = today - delta_t

        elif  ((type (stop_time) == float) | (type (stop_time) == int)):

                if ((stop_timedelta_unit == 'day') | (stop_timedelta_unit == 'd')):

                    UNIT = 'd'

                elif ((stop_timedelta_unit == 'hour') | (stop_timedelta_unit == 'h')):

                    UNIT = 'h'

                elif ((stop_timedelta_unit == 'minute') | (stop_timedelta_unit == 'm')):

                    UNIT = 'm'

                elif ((stop_timedelta_unit == 'second') | (stop_timedelta_unit == 's')):

                    UNIT = 's'

                elif ((stop_timedelta_unit == 'millisecond') | (stop_timedelta_unit == 'ms')):

                    UNIT = 'ms'

                delta_t = pd.Timedelta(stop_time, UNIT)
                STOP_TIMESTAMP = today - delta_t

        self.start_timestamp = START_TIMESTAMP
        self.stop_timestamp = STOP_TIMESTAMP
        
        # Convert to the the IP21 scale:
        self = self.convert_window_to_ip21_timescale()
        # Check if the conversion was performed correctly:
        print(f"Extracting entries from {self.start_timestamp} to {self.stop_timestamp}.")
        print(f"Start to Stop timestamp interval in IP21 timescale = ({self.start_ip21_scale}, {self.stop_ip21_scale})\n")
        
        return self
    
    
    def get_rest_api_url (self):
        
        server = self.server
        tag = self.tag
        data_source = self.data_source
        start_ip21_scale = self.start_ip21_scale
        stop_ip21_scale = self.stop_ip21_scale
        
        # URL Encodings:
        # https://docs.osisoft.com/bundle/pi-web-api-reference/page/help/topics/url-encoding.html
        
        # Check if the last character of the server is /:
        if (server[-1] != "/"):
            # Concatenate the character:
            server = server + "/"
        
        url = "http://" + server
        # To access the API Portal:
        # processdata_portal_suffix = "processdata/samples/sample_home.html/"
        # api_portal_url = url + processdata_portal_suffix
        query_prefix = "ProcessData/AtProcessDataREST.dll/History"
        # For the full URL:
        # ? indicates a query
        url = url + query_prefix
        
        # URL-encoded URL:
        query = f"""%3CQ%20f=%22d%22%20allQuotes=%221%22%3E%3CTag%3E%3CN%3E%3C![CDATA[{tag}]]%3E%3C/N%3E%3CD%3E%3C![CDATA[{data_source}]]%3E%3C/D%3E%3CF%3E%3C![CDATA[VAL]]%3E%3C/F%3E%3CHF%3E0%3C/HF%3E%3CSt%3E{start_ip21_scale}%3C/St%3E%3CEt%3E{stop_ip21_scale}%3C/Et%3E%3CRT%3E0%3C/RT%3E%3CX%3E100000%3C/X%3E%3CO%3E1%3C/O%3E%3C/Tag%3E%3C/Q%3E"""
        
        # Not-encoded URL:
        #query = f"""<Q%20f="d"%20allQuotes="1"><Tag><N><![CDATA[{tag}]]></N><D><![CDATA[{data_source}]]></D><F><![CDATA[VAL]]></F><HF>0</HF><St>{start_ip21_scale}</St><Et>{stop_ip21_scale}</Et><RT>0</RT><X>100000</X><O>1</O></Tag></Q>"""
        
        
        # Save the url as an attribute and return it:
        self.url = url
        self.query = query
        
        return self
    
    
    def retrieve_pd_dataframe (self, json_file_path = None):
        
        import os
        import json
        from pandas import json_normalize
        
        json_response = self.json_response
        tag_name = self.actual_tag_name
        start_ip21_scale = self.start_ip21_scale
        stop_ip21_scale = self.stop_ip21_scale
        
        # Retrieve previous dataset in memory:
        previous_df = self.dataset
        
        if (json_file_path is not None):
            try:
                # Extract the file extension
                file_extension = os.path.splitext(json_file_path)[1][1:]
                # os.path.splitext(file_path) is a tuple of strings: the first is the complete file
                # root with no extension; the second is the extension starting with a point: '.txt'
                # When we set os.path.splitext(file_path)[1], we are selecting the second element of
                # the tuple. By selecting os.path.splitext(file_path)[1][1:], we are taking this string
                # from the second character (index 1), eliminating the dot: 'txt'

                if (file_extension == 'json'):

                    json_file = json.load(json_response)

                else:
                    # Open context manager:
                    with open (json_response, 'r') as file:
                        # Read all lines:
                        response = file.readlines()

                    # Use the json.loads method to convert the string to json
                    json_file = json.loads(response)
            
            except:
                pass
        
        else:
            try:
                # It is a string:
                json_file = json.loads(json_response)
            
            except:
                pass
        
        """
        JSON structure obtained from IP21:
        {"data":[{...,"samples":[{"t":TIMESTAMP in IP21 scale,"v": VALUE FOR THAT TIMESTAMP,...},...]}]}
        
        """
        # https://pandas.pydata.org/pandas-docs/stable/reference/api/pandas.json_normalize.html#pandas.json_normalize
        # json_record_path (string): manipulate parameter 'record_path' from json_normalize method.
        # Path in each object to list of records. If not passed, data will be assumed to 
        # be an array of records. If a given field from the JSON stores a nested JSON (or a nested
        # dictionary) declare it here to decompose the content of the nested data. e.g. if the field
        # 'books' stores a nested JSON, declare, json_record_path = 'books'

        # json_field_separator = "_" (string). Manipulates the parameter 'sep' from json_normalize method.
        # Nested records will generate names separated by sep. 
        # e.g., for json_field_separator = ".", {‘foo’: {‘bar’: 0}} -> foo.bar.
        # Then, if a given field 'main_field' stores a nested JSON with fields 'field1', 'field2', ...
        # the name of the columns of the dataframe will be formed by concatenating 'main_field', the
        # separator, and the names of the nested fields: 'main_field_field1', 'main_field_field2',...

        # e.g. Suppose a JSON with the following structure: {'name': 'Mary', 'last': 'Shelley',
        # 'books': [{'title': 'Frankestein', 'year': 1818}, {'title': 'Mathilda ', 'year': 1819},{'title': 'The Last Man', 'year': 1826}]},
        # Here, there are nested JSONs in the field 'books'. The fields that are not nested
        # are 'name' and 'last'.
        # Then, json_record_path = 'books'
        # json_metadata_prefix_list = ['name', 'last']
        json_record_path = ['data', 'samples']
        json_field_separator = "_"
        dataset = json_normalize(json_file, record_path = json_record_path, sep = json_field_separator)
        
        if ((dataset is None) | ("er" in dataset.columns) | ("ec" in dataset.columns) | ("es" in dataset.columns) | (len(dataset) == 0)):
            
            print("There is no data available for the defined time window.\n")
            
            # If the attribute is not None, concatenate the dataframes:
            if (previous_df is not None):
                # Concatenate all dataframes (append rows):
                dataset = previous_df
                print("Returning the previous dataset itself.\n")
        
            self.dataset = dataset
            self.need_next_call = False
            # Interrupt the algorithm:
            return self
            
        
        if (('t' not in dataset.columns) & ('v' not in dataset.columns)):
            
            # Lower case column names:
            columns = [c.lower() for c in dataset.columns]
            # Pick only first character (until character of index 1, excluding index 1):
            columns = [c[:1] for c in columns]
            # Make this list the columns names:
            dataset.columns = columns
        
        if (('t' in dataset.columns) & ('v' in dataset.columns)):
            
            # Keep only columns 't' and 'v':
            dataset = dataset[['t', 'v']]
            # Rename these columns:
            dataset.columns = ['timestamp_ip21_scale', tag_name]
        
        else:
            # Substitute only the 1st column name and pick the other columns starting from
            # the second one (index 1)
            columns = ['timestamp_ip21_scale'] + list(dataset.columns)[1:]
        
        # Isolate the time series:
        time_series = dataset['timestamp_ip21_scale']
        # Get the last element (convert to list to access index -1).
        # In Pandas series, we can do [:-1] to get a single-element series:
        last_element = list(time_series)[-1]
        
        if (last_element < stop_ip21_scale):
            
            self.need_next_call = True
            # Update the start timestamp to be the last_element plus 1 unit (1 millisecond):
            self.start_ip21_scale = last_element + 1
        
        else:
            self.need_next_call = False
        
        # Save the timestamps as numpy array in the attribute ip21time_array:
        self.ip21time_array = np.array(dataset['timestamp_ip21_scale'])
        # Convert to timestamp using the conversion method:
        self = self.convert_ip21_timescale_array_to_timestamp()
        # Retrieve the timestamps:
        dataset['timestamp'] = self.timestamp_series
        
        # Select only these columns:
        dataset = dataset[['timestamp', tag_name]]
        
        
        # If the attribute is not None, concatenate the dataframes:
        if (previous_df is not None):
            
            # Compare the last timestamp from dataset and previous_df
            last_timestamp = list(dataset['timestamp'])[-1]
            previous_last_timestamp = list(previous_df['timestamp'])[-1]
            
            if (last_timestamp == previous_last_timestamp):
                
                # We already reached the end of the database. It actually never reaches the stop timestamp.
                print(f"The last timestamp registered in the database is: {last_timestamp}\n")
                # Pick the previous dataframe, which is already concatenated with dfs from previous calls
                dataset = previous_df
                # Now, finish the process:
                self.need_next_call = False
            
            else:
                # Concatenate all dataframes (append rows):
                dataset = pd.concat([previous_df, dataset], axis = 0, join = "inner")
                # Reset previous indices so that numeration is continuous:
                dataset = dataset.reset_index(drop = True)

        # Finally, save the concatenated dataset as dataset attribute and return the object:
        self.dataset = dataset
        
        return self
    

    def fetch_database (self, request_type = 'get'):
        
        import requests
        # IP21 uses the NTLM authentication protocol
        from requests_ntlm import HttpNtlmAuth
       
        url = self.url
        query = self.query
        username = self.username
        password = self.password
        
        # Create the NTLM Authorization object:
        AUTH = HttpNtlmAuth(username, password)
        
        # IP21 requires the 'post' protocol
        
        if (request_type == 'post'):
            
            json_response = requests.post(url, auth = AUTH, data = query)
        
        else: #get
            
            url = url + "?" + query
            json_response = requests.get(url, auth = AUTH)
        
        json_response = json_response.text
        
        self.json_response = json_response
        
        return self


class SQLServerConnection:
    """
    Class for extracting data from a SQL Server instance.
    def __init__ (self, server, 
                  database,
                  username = '', 
                  password = '',
                  system = 'windows'):
    
    : param: system = 'windows', 'macos' or 'linux'

        If the user passes the argument, use them. Otherwise, use the standard values.
        Set the class objects' attributes.
        Suppose the object is named assistant. We can access the attribute as:
        assistant.assistant_startup, for instance.
        So, we can save the variables as objects' attributes.

    INSTALL ODBC DRIVER IF USING MACOS OR LINUX (Windows already has it):
    - MacOS Installation: https://learn.microsoft.com/en-us/sql/connect/odbc/linux-mac/install-microsoft-odbc-driver-sql-server-macos?view=sql-server-ver16
    - Linux Installation: https://learn.microsoft.com/en-us/sql/connect/odbc/linux-mac/installing-the-microsoft-odbc-driver-for-sql-server?view=sql-server-ver16&tabs=alpine18-install%2Calpine17-install%2Cdebian8-install%2Credhat7-13-install%2Crhel7-offline
    """

    # Initialize instance attributes.
    # define the Class constructor, i.e., how are its objects:

    def __init__ (self, server, 
                  database,
                  username = '', 
                  password = '',
                  system = 'windows'):

        
        import pyodbc
        # Some other example server values are
        # server = 'localhost\sqlexpress' # for a named instance
        # server = 'myserver,port' # to specify an alternate port

        if ((username is None)|(username == '')):
            # Ask the user to provide the credentials:
            username = input(f"Enter your username for accessing the SQL database {database} from server {server} here (in the right).")
            print("\n") # line break
        
        if ((password is None)|(password == '')):
            from getpass import getpass
            password = getpass(f"Enter your password (Secret key) for accessing the SQL database {database} from server {server} here (in the right).")
        
        
        if (system == 'windows'):
            cnxn = pyodbc.connect('DRIVER={SQL Server};SERVER=' + server + ';DATABASE=' + database + ';UID=' + username + ';PWD=' + password)
        
        else:
            cnxn = pyodbc.connect('DRIVER={SQL Server};SERVER=' + server + ';DATABASE=' + database + ';UID=' + username + ';PWD=' + password + 'Encrypt=no;TrustServerCertificate=yes')
            # https://stackoverflow.com/questions/71587239/operationalerror-when-trying-to-connect-to-sql-server-database-using-pyodbc/71588236#71588236
        
        cursor = cnxn.cursor()
        
        self.cnxn = cnxn
        self.cursor = cursor
        self.query_counter = 0
        

    def get_db_schema (self, show_schema = True, export_csv = False, saving_directory_path = "db_schema.csv"):
        """
        : param: show_schema (bool): if True, the schema of the tables on the SQL Server will be shown.
        """
            
        query = "SELECT * FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_TYPE='BASE TABLE'"
            
        schema_df = pd.read_sql(query, self.cnxn)
        
        if ControlVars.show_results: # dominant context
            if (show_schema):
                print("Database schema:")
                print(f"There are {len(schema_df)} tables registered in the database.\n")
                try:
                    from IPython.display import display
                    display(schema_df)
                    
                except:
                    print(schema_df)
                
        self.schema_df = schema_df

        if (export_csv):
            if ((saving_directory_path is None)|(saving_directory_path == '')):
                saving_directory_path = "db_schema.csv"
            
            schema_df.to_csv(saving_directory_path)
            
        return self
        
        
    def run_sql_query (self, query, show_table = True, export_csv = False, saving_directory_path = ""):
        """
        : param: show_table (bool): keep as True to print the queried table, set False to hide it.
        : param: export_csv (bool): set True to export the queried table as CSV file, or set False not to export it.
        : param: saving_directory_path (str): full path containing directories and table name, 
            with .csv extension, used when export_csv = True
        """

        # SQL SERVER INNER JOIN:
        """
        https://www.sqlservertutorial.net/sql-server-basics/sql-server-inner-join/

        FROM table1 d
        INNER JOIN table2 t
        ON d.key1 = t.key2
        """
        
        # SQL SERVER UNION STATEMENT (vertical concatenation/append):
        """
        https://www.w3schools.com/sql/sql_union.asp

        FROM table1 d
        INNER JOIN table2 t
        ON d.key1 = t.key2
        UNION
        SELECT ValueTime as timestamp, TagName AS tag
        FROM LatestIP21TagDataNumeric
        WHERE TagName = 'TAGABC' OR TagName = 'TAGABD' OR TagName = 'TAGABE' 
        ORDER BY timestamp ASC;
        -- Notice that ORDER BY appears at the end of the query, after all joins and unions
        """

        # SQL SERVER TOP (EQUIVALENT TO LIMIT):
        """
        https://www.w3schools.com/sqL/sql_top.asp
        
        SELECT TOP 100 *
        FROM table;
        """
        
        # SQL SERVER WHERE STATEMENT (FILTER):
        """
        https://learn.microsoft.com/en-us/sql/t-sql/queries/where-transact-sql?view=sql-server-ver16

        WHERE column1 = 'val1' OR column1 = 'val2'
        """
        
        # SQL SERVER CASE STATEMENT (IF ELSE):
        """
        https://www.w3schools.com/sql/sql_case.asp

        CASE
            WHEN t.TagName = 'CODEXXX' THEN 'Variable X'
        ELSE ''
        END AS variable,
        """

        query_counter = self.query_counter
            
        df = pd.read_sql(query, self.cnxn)

        if ControlVars.show_results: # dominant context
            if (show_table):   
                print("Returned table:\n")
                try:
                    from IPython.display import display
                    display(df)
                    
                except:
                    print(df)

        # Vars function allows accessing the attributes from a class as a key from a dictionary.
        # vars(object) is a dictionary where each key is one attribute from the object. A new attribute may be
        # created by setting a value for a new key;
        # Then, an attribute name may be created as a string:
        vars(self)[f"df_query{query_counter}"] = df

        if (export_csv):
            if ((saving_directory_path is None)|(saving_directory_path == '')):
                saving_directory_path = f"table{query_counter}.csv"
            
            df.to_csv(saving_directory_path)
        
        # Update counter:
        self.query_counter = query_counter + 1

        return df
        
        
    def get_full_table (self, table, show_table = True, export_csv = False, saving_directory_path = ""):
        """
        : param: table (str): string containing the name of the table that will be queried.
        """
        
        query_counter = self.query_counter

        query = "SELECT * FROM " + str(table)
            
        df_table = pd.read_sql(query, self.cnxn)
        
        if ControlVars.show_results: # dominant context
            if (show_table): 
                print("Returned table:\n")
                try:
                    from IPython.display import display
                    display(df_table)
                    
                except:
                    print(df_table)
        
        # Vars function allows accessing the attributes from a class as a key from a dictionary.
        # vars(object) is a dictionary where each key is one attribute from the object. A new attribute may be
        # created by setting a value for a new key;
        # Then, an attribute name may be created as a string:
        vars(self)[f"df_table{query_counter}"] = df_table

        if (export_csv):
            if ((saving_directory_path is None)|(saving_directory_path == '')):
                saving_directory_path = f"table{query_counter}.csv"
            
            df_table.to_csv(saving_directory_path)
         
        # Update counter:
        self.query_counter = query_counter + 1    
            
        return df_table
    
    
    def query_specific_tag_ip21sqlserver (self, tag, variable_name = None, show_table = True, export_csv = False, saving_directory_path = ""):
        """ : param: tag (str): string with tag as registered in IP21. e.g. tag = 'ABC00AA101-01'.
        
            : param: variable_name (str): string containing a more readable name for the tag, that will be also shown.
            e.g. variable_name = 'Temperature in C'
        """
        
        # https://www.sqlservertutorial.net/sql-server-basics/sql-server-inner-join/
        # https://www.w3schools.com/sqL/sql_top.asp
        # https://learn.microsoft.com/en-us/sql/t-sql/queries/where-transact-sql?view=sql-server-ver16
        # https://www.w3schools.com/sql/sql_case.asp
        # https://www.w3schools.com/sql/sql_union.asp
        
        if (variable_name is None):
            #Repeat the tag
            variable_name = tag

        query = f"""SELECT d.ValueTime AS timestamp, t.TagName AS tag,
                    CASE
                        WHEN t.TagName = '{tag}' THEN '{variable_name}'
                        ELSE ''
                    END AS variable,
                    d.Value AS value
                    FROM IP21DataNumeric d
                    INNER JOIN IP21PublishConfig t
                    ON d.TagConfigID = t.ID
                    WHERE t.TagName = '{tag}'
                    UNION
                    SELECT ValueTime as timestamp, TagName AS tag, 
                    CASE
                        WHEN TagName = '{tag}' THEN '{variable_name}'
                        ELSE ''
                    END AS variable,
                    Value AS value
                    FROM LatestIP21TagDataNumeric
                    WHERE TagName = '{tag}'
                    ORDER BY timestamp ASC;
                """
        
        tag_df = self.run_sql_query(query = query, show_table = show_table, export_csv = export_csv, saving_directory_path = saving_directory_path)
        
            
        return tag_df


class SQLiteConnection:
    """Class for connecting and manipulating a SQLite Database file"""

    def __init__(self, file_path, pre_created_engine = None):

        self.file_path = file_path
        self.engine = pre_created_engine
    

    def create_engine(self):
        
        # Make imports and create the engine for the database
        # Configure the SQLite engine
        from sqlalchemy import create_engine
        
        # SQLAlchemy engines documentation
        # https://docs.sqlalchemy.org/en/20/core/engines.html
        # SQLite connects to file-based databases, using the Python built-in module sqlite3 by default.
        # As SQLite connects to local files, the URL format is slightly different. 
        # The “file” portion of the URL is the filename of the database. For a relative file path, this requires 
        # three slashes:
        # sqlite://<nohostname>/<path>
        # where <path> is relative:

        try:
                    
            if (self.file_path[:2] == './'):
                # Add a slash, since sqlite engine requires 3 slashes
                self.file_path = self.file_path[2:]
                
            if (self.file_path[0] == '/'):
                # Add a slash, since sqlite engine requires 3 slashes
                self.file_path = self.file_path[1:]
                        
            self.file_path = """sqlite:///""" + self.file_path
            # file_path = "sqlite:///my_db.db"
                    
            self.engine = create_engine(self.file_path)
            #And for an absolute file path, the three slashes are followed by the absolute path:
                
            """
            # Unix/Mac - 4 initial slashes in total
            engine = create_engine("sqlite:////absolute/path/to/foo.db")
                
            # Windows
            engine = create_engine("sqlite:///C:\\path\\to\\foo.db")
                
            # Windows alternative using raw string
            engine = create_engine(r"sqlite:///C:\path\to\foo.db")
            To use a SQLite :memory: database, specify an empty URL:
                
            engine = create_engine("sqlite://")
            More notes on connecting to SQLite at SQLite.
            """
        
        except:
            raise InvalidInputsError ("Error trying to create SQLite Engine Database. Check if no more than one slash was added to file path.\n")
        
        return self
    

    def fetch_table(self, table_name):
        
        # If there is no engine, create one:
        if (self.engine is None):
            self = self.create_engine()
        
        try:
            # Access the table from the database
            df = pd.read_sql(table_name, self.engine)
            
            if ControlVars.show_results: 
                print(f"Successfully retrieved table {table_name} from the database.")
                print("Check the 10 first rows of the dataframe:\n")
                
                try:
                    # only works in Jupyter Notebook:
                    from IPython.display import display
                    display(df.head(10))
                        
                except: # regular mode
                    print(df.head(10))
            
            return df, self.engine
        
        except:
            raise InvalidInputsError ("Error trying to fetch SQLite Database. If an pre-created engine was provided, check if it is correct and working.\n")
        

    def update_or_create_table(self, table_name):
    
        # If there is no engine, create one:
        if (self.engine is None):
            self = self.create_engine()
            
        try:
            # Set index = False not to add extra indices in the database:
            df.to_sql(table_name, con = engine, if_exists = 'replace', index = False)
                
            if ControlVars.show_results: 
                print(f"Successfully updated table {table_name} on the SQLite database.")
                print("Check the 10 first rows from this table:\n")
                    
                try:
                    # only works in Jupyter Notebook:
                    from IPython.display import display
                    display(df.head(10))
                            
                except: # regular mode
                    print(df.head(10))

            return df, self.engine
        
        except:
            raise InvalidInputsError ("Error trying to update SQLite Database. If an pre-created engine was provided, check if it is correct and working.\n")
            

class GCPBigQueryConnection:
    """
    Class for accessing Google Cloud Platform (GCP) BigQuery data.

    : param: project (str): project name on BigQuery
    : param: dataset (str): dataset that user wants to connect

    : param: already_authenticated (bool): True if the manual connection to GCP was already performed
        and so the user is authorized.
        
        This connection can be done by running the following command directly on a cell or on the console
        from Python environment. Attention: follow the command line authentication instruction below.

            !gcloud auth application-default login
        
    
        COMMAND LINE AUTHENTICATION INSTRUCTION
            
            - BEFORE INSTATIATING THE CLASS, FOLLOW THESE GUIDELINES!
            
            1. Copy and run the following line in a notebook cell. Do not add quotes.
            The "!" sign must be added to indicate the use of a command line software:

                !gcloud auth application-default login
            
            2. Also, you can run the SQL query on GCP console and, when the query results appear, click on EXPLORE DATA - Explore with Python notebook.
            It will launch a Google Colab Python notebook that you may run for data exploring, manipulation and exporting.

            2.1. To export data, you expand the hamburger icon on the left side of the Google Colab, click on Files, and then select an exported CSV or other files. Finally, click on the ellipsis (3 dots) and select Download to obtain it.
            

        Install Google Cloud Software Development Kit (SDK) before running
                
        General Instructions for Installation: https://cloud.google.com/sdk/docs/install-sdk?hl=pt-br#installing_the_latest_version
        Instructions for Windows: https://cloud.google.com/sdk/docs/install-sdk?hl=pt-br#windows
        Instructions for Mac OS: https://cloud.google.com/sdk/docs/install-sdk?hl=pt-br#mac
        Instructions for Linux: https://cloud.google.com/sdk/docs/install-sdk?hl=pt-br#linux
        
        From: https://stackoverflow.com/questions/39419754/downloading-and-importing-google-cloud-python
        First, make sure you have installed gcloud on your system then run the commands like this:

        First: gcloud components update in your terminal.
        then: pip install google-cloud
        And for the import error:
        Adding "--ignore-installed" to pip command may work.
        This might be a bug in pip - see this page for more details: https://github.com/pypa/pip/issues/2751

        This pipeline may be blocked also due to security configurations, and may fail on some Virtual Private Networks (VPNs).
        - Try to run this pipeline outside of the VPN in case it fails

    """
    
    # Initialize instance attributes.
    # define the Class constructor, i.e., how are its objects:

    def __init__ (self, project = '', dataset = '', already_authenticated = True):
       

        from google.cloud import bigquery
        from google.cloud import bigquery_storage
        from google.oauth2 import service_account
        import google.auth
        
        if ((project is None)|(project == '')):
            # Ask the user to provide the credentials:
            # This is the name that appears on the right-hand menu from GCP console Big Query page, 
            # (https://console.cloud.google.com/bigquery) called Explorer
            # that contains a group of datasets. E.g.: location360-datasets; bcs-csw-core, etc
            # The individual datasets are revealed after expanding the project name by clicking on the arrow.
            print("\n")
            self.project = input(f"Enter the name of the project registered on Google Cloud Platform (GCP).\n")
            
        if ((dataset is None)|(dataset == '')):
            # Ask the user to provide the credentials:
            # E.g.: core
            print("\n")
            self.dataset = input(f"Enter the name of the dataset from project {self.project} registered on GCP, containing the tables that will be queried.\n")

        self.query_counter = 0
        self.already_authenticated = already_authenticated
        

    def authenticate (self, authentication_method = 'manual',
                  vault_secret_path = '', app_role = '', app_secret = ''):
        """
        : param: authentication_method (str): 'manual' or 'vault' authentication
            : param: authetication_method = 'manual' for GCP standard manual authentication on browser. System will try
                to access the authorization window, in case it was not done yet.
            : param: authetication_method = 'vault' for Vault automatic authentication, dependent on corporate
                cibersecurity and data asset.

        : param: vault_secret_path (str): path to access the secret
        : params: vault_secret_path = '', app_role = '', app_secret = '' are the parameters for vault authorization
            
        """
        
        if self.already_authenticated:
            self.bqclient = bigquery.Client(project = self.project)
            try:
                self.bqstorageclient = bigquery_storage.BigQueryReadClient()
            except:
                self.bqstorageclient = None # create an empty attribute
        
        else:

            if (authentication_method == 'manual'):
                self = self.manual_authentication()
            
            elif (authentication_method == 'vault'):
                self = self.vault_authentication(vault_secret_path = vault_secret_path, app_role = app_role, app_secret = app_secret)

        return self


    def manual_authentication (self):

        from google.cloud import bigquery
        from google.cloud import bigquery_storage
        from google.oauth2 import service_account
        import google.auth
        
        try:
            # Setting authorization from cloud. Might seem no effect if CLI is configured
            self.credentials, self.project = google.auth.default(
                scopes = ["https://www.googleapis.com/auth/cloud-platform"]) 
            
            self.bqclient = bigquery.Client(credentials = self.credentials, project = self.project)
            self.bqstorageclient = bigquery_storage.BigQueryReadClient(credentials = self.credentials)
        
        except:
            # Some other example server values are
            # server = 'localhost\sqlexpress' # for a named instance
            # server = 'myserver,port' # to specify an alternate port
            
            # subprocess module allows running shell commands. Each portion from a Bash script is declared as an element 
            # from a list of strings. Outputs are captured as a list of strings as well.        
            # Example:
            """
            with Popen(["ls"], stdout=PIPE) as proc:
            out = proc.readlines()
            print(out)
            
            output: ['some_file.txt','some_other_file.txt']
            # Notice that a Python variable of string type may be included to the list for generating dynamic execution of commands.
            """

            """
            Example 2 command: python -m pip show pandas
            from subprocess import Popen, PIPE, TimeoutExpired
            proc = Popen(["python", "-m", "pip", "show", "pandas"], stdout = PIPE, stderr = PIPE)
            try:
                output, error = proc.communicate(timeout = 15)
                print(output)
            except:
                        # General exception
                output, error = proc.communicate()
                print(f"Process with output: {output}, error: {error}.\n")

            output: b'Name: pandas\r\nVersion: 2.0.3\r\nSummary: Powerful data structures for data analysis, time series, and statistics\r\nHome-page: \r\nAuthor: \r\nAuthor-email: The Pandas Development Team <pandas-dev@python.org>\r\nLicense: BSD 3-Clause License\r\n        \r\n        Copyright (c) 2008-2011, AQR Capital Management, LLC, Lambda Foundry, Inc. and PyData Development Team\r\n        All rights reserved.\r\n        \r\n        Copyright (c) 2011-2023, Open source contributors.\r\n        \r\n        Redistribution and use in source and binary forms, with or without\r\n        modification, are permitted provided that the following conditions are met:\r\n        \r\n        * Redistributions of source code must retain the above copyright notice, this\r\n          list of conditions and the following disclaimer.\r\n        \r\n        * Redistributions in binary form must reproduce the above copyright notice,\r\n          this list of conditions and the following disclaimer in the documentation\r\n          and/or other materials provided with the distribution.\r\n        \r\n        * Neither the name of the copyright holder nor the names of its\r\n          contributors may be used to endorse or promote products derived from\r\n          this software without specific prior written permission.\r\n        \r\n        THIS SOFTWARE IS PROVIDED BY THE COPYRIGHT HOLDERS AND CONTRIBUTORS "AS IS"\r\n        AND ANY EXPRESS OR IMPLIED WARRANTIES, INCLUDING, BUT NOT LIMITED TO, THE\r\n        IMPLIED WARRANTIES OF MERCHANTABILITY AND FITNESS FOR A PARTICULAR PURPOSE ARE\r\n        DISCLAIMED. IN NO EVENT SHALL THE COPYRIGHT HOLDER OR CONTRIBUTORS BE LIABLE\r\n        FOR ANY DIRECT, INDIRECT, INCIDENTAL, SPECIAL, EXEMPLARY, OR CONSEQUENTIAL\r\n        DAMAGES (INCLUDING, BUT NOT LIMITED TO, PROCUREMENT OF SUBSTITUTE GOODS OR\r\n        SERVICES; LOSS OF USE, DATA, OR PROFITS; OR BUSINESS INTERRUPTION) HOWEVER\r\n        CAUSED AND ON ANY THEORY OF LIABILITY, WHETHER IN CONTRACT, STRICT LIABILITY,\r\n        OR TORT (INCLUDING NEGLIGENCE OR OTHERWISE) ARISING IN ANY WAY OUT OF THE USE\r\n        OF THIS SOFTWARE, EVEN IF ADVISED OF THE POSSIBILITY OF SUCH DAMAGE.\r\n        \r\nLocation: c:\\users\\gnklm\\appdata\\local\\anaconda3\\lib\\site-packages\r\nRequires: numpy, python-dateutil, pytz, tzdata\r\nRequired-by: cmdstanpy, datashader, holoviews, hvplot, prophet, seaborn, shap, statsmodels, xarray\r\n'
            """

            from subprocess import Popen, PIPE, TimeoutExpired

            # Start a long running process using subprocess.Popen()
            # Command to run:
            """!gcloud auth application-default login"""

            try:
                proc = Popen(["gcloud", "auth", "application-default", "login"], stdout = PIPE, stderr = PIPE)

                """
                You will use the subprocess.communicate() method to wait for the command to finish running for up to 15 seconds. 
                The process will then timeout and it will return an Exception: i.e. error detected during execution, which will be caught 
                and the process will be cleaned up by proc.kill(). 
                """

                # Use subprocess.communicate() to create a timeout 
                try:
                    output, error = proc.communicate(timeout = 15)
                    # Simply remove timeout argument if process is not supposed to finish after a given time.

                except TimeoutExpired:

                    # Cleanup the process if it takes longer than the timeout
                    proc.kill()

                    # Read standard out and standard error streams and print
                    output, error = proc.communicate()
                    print(f"Process timed out with output: {output}, error: {error}.")

                except:
                    # General exception
                    proc.kill()
                    output, error = proc.communicate()
                    print(f"Process of Google Cloud Mount with output: {output}, error: {error}.\n")

                    warning = """
                    Install Google Cloud Software Development Kit (SDK) before running this function.

                    General Instructions for Installation: https://cloud.google.com/sdk/docs/install-sdk?hl=pt-br#installing_the_latest_version
                    Instructions for Windows: https://cloud.google.com/sdk/docs/install-sdk?hl=pt-br#windows
                    Instructions for Mac OS: https://cloud.google.com/sdk/docs/install-sdk?hl=pt-br#mac
                    Instructions for Linux: https://cloud.google.com/sdk/docs/install-sdk?hl=pt-br#linux
                    """

                    print(warning)

                """try:
                    from IPython.display import display 
                    # from IPython.display import display, display_html
                    out = list(proc._fileobj2output.keys())
                    out1, out2 = out[1], out[0]
                    display(proc._fileobj2output[out1][0])
                    print("\n")
                    # display(proc._fileobj2output[out2][0])
                    # display_html(proc._fileobj2output[out2][0])

                except:
                    pass"""
            
            except:
                pass
        
            try:
                self.bqclient = bigquery.Client(project = self.project)
            except:

                error_msg = """
                Impossible to Connect with the input parameters or protocols. 
                Try manual connection running

                    !gcloud auth application-default login
                
                on a cell or on the Python IDE's terminal.
                """
                raise InvalidInputsError(error_msg)


        return self


    def get_vault_secret (self, vault_secret_path: str, app_role: str, app_secret: str):
               
        import base64
        import json
        import hvac
        import os

        from google.oauth2 import service_account
        from google.cloud import bigquery, bigquery_storage
        
        """
        Get the vault secret in dictionary
        More detail about hvac: https://hvac.readthedocs.io/en/stable/overview.html
        
        """
        vault_url = 'https://vault.agro.services'
        vault_client = hvac.Client(url = vault_url)
        
        vault_client.auth.approle.login(app_role, app_secret)

        vault_path = vault_secret_path
        vault_secret = vault_client.read(vault_path)
        
        return vault_secret['data']['data']

 
    def get_vault_credentials (self, vault_secret_path: str, app_role: str, app_secret: str):
        
        import base64
        import json
        import hvac
        import os

        from google.oauth2 import service_account
        from google.cloud import bigquery, bigquery_storage
        
        """
        Bigquery project credential with service account
        Return credential of a service account using app role and app secret
        
        """
        vault_secret = self.get_vault_secret(vault_secret_path, app_role, app_secret)

        if 'data' in vault_secret and type(vault_secret['data']) == str:
            service_account_creds = json.loads(base64.b64decode(vault_secret['data']))
        
        else:
            # in case credentials are saved directly as json object in vault (not encoded) you can get it directly
            service_account_creds = json.loads(base64.b64decode(vault_secret))
        
        bq_credentials = service_account.Credentials.from_service_account_info(service_account_creds)

        return bq_credentials


    def vault_authentication (self, vault_secret_path = '', app_role = '', app_secret = ''):     
        
        import base64
        import json
        import hvac
        import os

        from google.oauth2 import service_account
        from google.cloud import bigquery, bigquery_storage

        
        if ((vault_secret_path is None)|(vault_secret_path == '')):
            vault_secret_path = input(f"Enter Vault Secret Path for accessing CSW.")
            print("\n") # line break
        
        if ((app_role is None)|(app_role == '')):
            app_role = input(f"Enter App Role for getting the Vault client.")
            print("\n") # line break

        if ((app_secret is None)|(app_secret == '')):
            from getpass import getpass
            vault_secret_path = getpass(f"Enter App Secret for getting the Vault client.")
            print("\n") # line break

        self.credentials = self.get_vault_credentials(vault_secret_path = vault_secret_path, app_role = app_role, app_secret = app_secret)

        self.bqclient = bigquery.Client(credentials = self.credentials, project = self.project)
        self.bqstorageclient = bigquery_storage.BigQueryReadClient(credentials = self.credentials)

        return self


    def table_exists (self, table_id: str) -> bool:
        """
        Checks if a table with the specified ID exists in BigQuery.

        Parameters:
        - client (bigquery.Client): The BigQuery client.
        - table_id (str): The ID of the table to check.

        Returns:
        bool: True if the table exists, False otherwise.
        """
        client = self.bqclient

        try:
            client.get_table(table_id)
            return True
        
        except:
            return False


    def run_sql_query (self, query, show_table = True, export_csv = False, saving_directory_path = ""):
        """
        : param: show_table (bool): keep as True to print the queried table, set False to hide it.
        : param: export_csv (bool): set True to export the queried table as CSV file, or set False not to export it.
        : param: saving_directory_path (str): full path containing directories and table name, 
            with .csv extension, used when export_csv = True
        """

        query_counter = self.query_counter

        client = self.bqclient
        job = client.query(query)
        df = job.to_dataframe()

        if ControlVars.show_results: # dominant context
            if (show_table):   
                print("Returned table:\n")
                try:
                    from IPython.display import display
                    display(df)
                    
                except:
                    print(df)

        # Vars function allows accessing the attributes from a class as a key from a dictionary.
        # vars(object) is a dictionary where each key is one attribute from the object. A new attribute may be
        # created by setting a value for a new key;
        # Then, an attribute name may be created as a string:
        vars(self)[f"df_query{query_counter}"] = df

        if (export_csv):
            if ((saving_directory_path is None)|(saving_directory_path == '')):
                saving_directory_path = f"table{query_counter}.csv"
            
            df.to_csv(saving_directory_path)
        
        # Update counter:
        self.query_counter = query_counter + 1

        return df
        
        
    def get_full_table (self, table, show_table = True, export_csv = False, saving_directory_path = ""):
        """
        : param: table (str): name of the table to be retrieved. Full table name is `{self.project}.{self.dataset}.{str(table)}`
        : param: show_table (bool): keep as True to print the queried table, set False to hide it.
        : param: export_csv (bool): set True to export the queried table as CSV file, or set False not to export it.
        : param: saving_directory_path (str): full path containing directories and table name, 
            with .csv extension, used when export_csv = True
        """

        table_name = f"""`{self.project}.{self.dataset}.{str(table)}`"""
        
        query_counter = self.query_counter

        query = "SELECT * FROM " + table_name
            
        client = self.bqclient
        job = client.query(query)
        df_table = job.to_dataframe()
        
        if ControlVars.show_results: # dominant context
            if (show_table): 
                print("Returned table:\n")
                try:
                    from IPython.display import display
                    display(df_table)
                    
                except:
                    print(df_table)
        
        # Vars function allows accessing the attributes from a class as a key from a dictionary.
        # vars(object) is a dictionary where each key is one attribute from the object. A new attribute may be
        # created by setting a value for a new key;
        # Then, an attribute name may be created as a string:
        vars(self)[f"df_table{query_counter}"] = df_table

        if (export_csv):
            if ((saving_directory_path is None)|(saving_directory_path == '')):
                saving_directory_path = f"table{query_counter}.csv"
            
            df_table.to_csv(saving_directory_path)
         
        # Update counter:
        self.query_counter = query_counter + 1    
            
        return df_table
    

    def write_data_on_bigquery_table (self, table, df):
        """
        : param: table (str): string with table name
        : param: df (pd.DataFrame): Pandas dataframe to be written on BigQuery table
        """

        client = self.bqclient
        table_ref = client.dataset(self.dataset).table(str(table))
        table = client.get_table(table_ref)

        errors = client.insert_rows_from_dataframe(table, df)

        if errors != [[]]: 
            raise RuntimeError('Error in writing to BigQuery: {}'.format(errors))

        else:
            
            print(f"Dataframe written on Big Query {table} table from dataset {self.project}.{self.dataset}.\n")
            
            return self


    def delete_specific_values_from_column_on_table (self, table, column, values_to_delete, show_table = True, export_csv = False, saving_directory_path = ""):
        """
        : param: column (str): is the column name on a given BigQuery table (a string).
        : param: values_to_delete is a single value (numeric or string) or an iterable containing a set
          of values to be deleted.
        : param: show_table (bool): keep as True to print the queried table, set False to hide it.
        : param: export_csv (bool): set True to export the queried table as CSV file, or set False not to export it.
        : param: saving_directory_path (str): full path containing directories and table name, 
            with .csv extension, used when export_csv = True
        """
        
        # column is the column name on a given BigQuery table (a string).
        # values_to_delete is a single value (numeric or string) or an iterable containing a set
        # of values to be deleted.


        if (type(values_to_delete) == str):
            # put inside list:
            values_to_delete = [values_to_delete]
        
        else:
            try:
                values_to_delete = list(values_to_delete)
            
            except:
                # It is not an iterable (it is a value):
                values_to_delete = [values_to_delete]

        # pick 1st element and remove it from list
        val0 = values_to_delete.pop(0)

        if (type(val0) == str):
            
            delete_query = f"""
                                DELETE 
                                    FROM `{self.project}.{self.dataset}.{str(table)}`
                                    WHERE {column} = '{val0}' """
        
        else:
            delete_query = f"""
                                DELETE 
                                    FROM `{self.project}.{self.dataset}.{str(table)}`
                                    WHERE {column} = {val0} """

        # Now, loop through the remaining list (if there is a remaining one) to update query.
        # If list is empty, the loop does not run
        for val in values_to_delete:
            if (type(val) == str):
                delete_query = delete_query + f"""OR {column} = '{val}' """
                
            else:
                delete_query = delete_query + f"""OR {column} = {val} """

        client = self.bqclient
        query_counter = self.query_counter

        table_ref = client.dataset(self.dataset).table(str(table))
        table = client.get_table(table_ref)

        table.streaming_buffer

        if table.streaming_buffer is None:
            job = client.query(delete_query)
            job.result()
        else:
            raise RuntimeError("Table contains data in a streaming buffer, which cannot be updated or deleted. Please perform this action when the streaming buffer is empty (may take up to 90 minutes from last data insertion).")

        df_table = job.to_dataframe()
        
        if ControlVars.show_results: # dominant context
            if (show_table): 
                print("Returned table with deleted data:\n")
                try:
                    from IPython.display import display
                    display(df_table)
                    
                except:
                    print(df_table)
        
        # Vars function allows accessing the attributes from a class as a key from a dictionary.
        # vars(object) is a dictionary where each key is one attribute from the object. A new attribute may be
        # created by setting a value for a new key;
        # Then, an attribute name may be created as a string:
        vars(self)[f"df_table{query_counter}"] = df_table

        if (export_csv):
            if ((saving_directory_path is None)|(saving_directory_path == '')):
                saving_directory_path = f"table{query_counter}.csv"
            
            df_table.to_csv(saving_directory_path)
         
        # Update counter:
        self.query_counter = query_counter + 1    
            
        return df_table    


    def update_specific_value_from_column_on_table (self, table, column, old_value, updated_value, show_table = True, export_csv = False, saving_directory_path = ""):
        """
        : param: column (str): is the column name on a given BigQuery table (a string).
        : param: old_value: value that must be replaced
        : param: updated_value: new value to be added.
        : param: show_table (bool): keep as True to print the queried table, set False to hide it.
        : param: export_csv (bool): set True to export the queried table as CSV file, or set False not to export it.
        : param: saving_directory_path (str): full path containing directories and table name, 
            with .csv extension, used when export_csv = True
        """
            
        if ((type(old_value) == str)|(type(updated_value) == str)):
                update_query = f"""
                    UPDATE `{self.project}.{self.dataset}.{str(table)}`
                        SET `{column}` = '{updated_value}'
                        WHERE `{column}` = '{old_value}'
                        """
            
        else:
                update_query = f"""
                    UPDATE `{self.project}.{self.dataset}.{str(table)}`
                        SET `{column}` = {updated_value}
                        WHERE `{column}` = {old_value}
                        """
            
        client = self.bqclient
        query_counter = self.query_counter

        table_ref = client.dataset(self.dataset).table(str(table))
        table = client.get_table(table_ref)

        table.streaming_buffer

        if table.streaming_buffer is None:
            job = client.query(update_query)
            job.result()
        else:
            raise RuntimeError("Table contains data in a streaming buffer, which cannot be updated or deleted. Please perform this action when the streaming buffer is empty (may take up to 90 minutes from last data insertion).")

        df_table = job.to_dataframe()
        
        if ControlVars.show_results: # dominant context
            if (show_table): 
                print("Returned table with updated data:\n")
                try:
                    from IPython.display import display
                    display(df_table)
                    
                except:
                    print(df_table)
        
        # Vars function allows accessing the attributes from a class as a key from a dictionary.
        # vars(object) is a dictionary where each key is one attribute from the object. A new attribute may be
        # created by setting a value for a new key;
        # Then, an attribute name may be created as a string:
        vars(self)[f"df_table{query_counter}"] = df_table

        if (export_csv):
            if ((saving_directory_path is None)|(saving_directory_path == '')):
                saving_directory_path = f"table{query_counter}.csv"
            
            df_table.to_csv(saving_directory_path)
         
        # Update counter:
        self.query_counter = query_counter + 1    
            
        return df_table


    def update_entire_column_from_table (self, table, column, updated_value, show_table = True, export_csv = False, saving_directory_path = ""):
        """
        : param: column (str): is the column name on a given BigQuery table (a string).
        : param: updated_value: new value to be added.
        : param: show_table (bool): keep as True to print the queried table, set False to hide it.
        : param: export_csv (bool): set True to export the queried table as CSV file, or set False not to export it.
        : param: saving_directory_path (str): full path containing directories and table name, 
            with .csv extension, used when export_csv = True
        """
            
        if (type(updated_value) == str):
                update_query = f"""
                    UPDATE `{self.project}.{self.dataset}.{str(table)}`
                        SET `{column}` = '{updated_value}'
                        WHERE TRUE
                        """
            
        else:
                update_query = f"""
                    UPDATE `{self.project}.{self.dataset}.{str(table)}`
                        SET `{column}` = {updated_value}
                        WHERE TRUE
                        """
            
        client = self.bqclient
        query_counter = self.query_counter

        table_ref = client.dataset(self.dataset).table(str(table))
        table = client.get_table(table_ref)

        table.streaming_buffer

        if table.streaming_buffer is None:
            job = client.query(update_query)
            job.result()
        else:
            raise RuntimeError("Table contains data in a streaming buffer, which cannot be updated or deleted. Please perform this action when the streaming buffer is empty (may take up to 90 minutes from last data insertion).")

        df_table = job.to_dataframe()
        
        if ControlVars.show_results: # dominant context
            if (show_table): 
                print("Returned table with updated data:\n")
                try:
                    from IPython.display import display
                    display(df_table)
                    
                except:
                    print(df_table)
        
        # Vars function allows accessing the attributes from a class as a key from a dictionary.
        # vars(object) is a dictionary where each key is one attribute from the object. A new attribute may be
        # created by setting a value for a new key;
        # Then, an attribute name may be created as a string:
        vars(self)[f"df_table{query_counter}"] = df_table

        if (export_csv):
            if ((saving_directory_path is None)|(saving_directory_path == '')):
                saving_directory_path = f"table{query_counter}.csv"
            
            df_table.to_csv(saving_directory_path)
         
        # Update counter:
        self.query_counter = query_counter + 1    
            
        return df_table    


    def update_value_when_finding_str_or_substring_on_another_column (self, table, column, updated_value, string_column, str_or_substring_to_search, show_table = True, export_csv = False, saving_directory_path = ""):
        """
        : param: column (str): is the column name on a given BigQuery table (a string).
        : param: updated_value: new value to be added.
        : param: string_column (str): column containing a string or substring that will be searched.
        : param: str_or_substring_to_search (str): (in quotes): string or substring that will be searched on 
            column 'string_column'. When it is find, the value on 'column' will be updated.
        : param: show_table (bool): keep as True to print the queried table, set False to hide it.
        : param: export_csv (bool): set True to export the queried table as CSV file, or set False not to export it.
        : param: saving_directory_path (str): full path containing directories and table name, 
            with .csv extension, used when export_csv = True
        """
        
        if (type(updated_value) == str):
                update_query = f"""
                    UPDATE `{self.project}.{self.dataset}.{str(table)}`
                        SET `{column}` = '{updated_value}'
                        WHERE CONTAINS_SUBSTR({string_column}, "{str_or_substring_to_search}")
                        """
            
        else:
                update_query = f"""
                    UPDATE `{self.project}.{self.dataset}.{str(table)}`
                        SET `{column}` = {updated_value}
                        WHERE CONTAINS_SUBSTR({string_column}, "{str_or_substring_to_search}")
                        """
            
        client = self.bqclient
        query_counter = self.query_counter

        table_ref = client.dataset(self.dataset).table(str(table))
        table = client.get_table(table_ref)

        table.streaming_buffer

        if table.streaming_buffer is None:
            job = client.query(update_query)
            job.result()
        else:
            raise RuntimeError("Table contains data in a streaming buffer, which cannot be updated or deleted. Please perform this action when the streaming buffer is empty (may take up to 90 minutes from last data insertion).")

        df_table = job.to_dataframe()
        
        if ControlVars.show_results: # dominant context
            if (show_table): 
                print("Returned table with updated data:\n")
                try:
                    from IPython.display import display
                    display(df_table)
                    
                except:
                    print(df_table)
            
        # Vars function allows accessing the attributes from a class as a key from a dictionary.
        # vars(object) is a dictionary where each key is one attribute from the object. A new attribute may be
        # created by setting a value for a new key;
        # Then, an attribute name may be created as a string:
        vars(self)[f"df_table{query_counter}"] = df_table

        if (export_csv):
            if ((saving_directory_path is None)|(saving_directory_path == '')):
                saving_directory_path = f"table{query_counter}.csv"
            
            df_table.to_csv(saving_directory_path)
         
        # Update counter:
        self.query_counter = query_counter + 1    
            
        return df_table


    def update_value_when_finding_numeric_value_on_another_column (self, table, column, updated_value, comparative_column, value_to_search, show_table = True, export_csv = False, saving_directory_path = ""):
        """
        : param: column (str): is the column name on a given BigQuery table (a string).
        : param: updated_value: new value to be added.
        : param: comparative_column (str): column containing a numeric value that will be searched.
        : param: value_to_search: numeric value that will be searched 
            on column 'comparative_colum'. When it is find, the value on 'column' will be updated.
        : param: show_table (bool): keep as True to print the queried table, set False to hide it.
        : param: export_csv (bool): set True to export the queried table as CSV file, or set False not to export it.
        : param: saving_directory_path (str): full path containing directories and table name, 
            with .csv extension, used when export_csv = True
        """
        
        if (type(updated_value) == str):
                update_query = f"""
                    UPDATE `{self.project}.{self.dataset}.{str(table)}`
                        SET `{column}` = '{updated_value}'
                        WHERE `{comparative_column}` = {value_to_search}
                        """
            
        else:
                update_query = f"""
                    UPDATE `{self.project}.{self.dataset}.{str(table)}`
                        SET `{column}` = {updated_value}
                        WHERE `{comparative_column}` = {value_to_search}
                        """
            
        client = self.bqclient
        query_counter = self.query_counter

        table_ref = client.dataset(self.dataset).table(str(table))
        table = client.get_table(table_ref)

        table.streaming_buffer

        if table.streaming_buffer is None:
            job = client.query(update_query)
            job.result()
        else:
            raise RuntimeError("Table contains data in a streaming buffer, which cannot be updated or deleted. Please perform this action when the streaming buffer is empty (may take up to 90 minutes from last data insertion).")

        df_table = job.to_dataframe()
        
        if ControlVars.show_results: # dominant context
            if (show_table): 
                print("Returned table with updated data:\n")
                try:
                    from IPython.display import display
                    display(df_table)
                    
                except:
                    print(df_table)
        
        # Vars function allows accessing the attributes from a class as a key from a dictionary.
        # vars(object) is a dictionary where each key is one attribute from the object. A new attribute may be
        # created by setting a value for a new key;
        # Then, an attribute name may be created as a string:
        vars(self)[f"df_table{query_counter}"] = df_table

        if (export_csv):
            if ((saving_directory_path is None)|(saving_directory_path == '')):
                saving_directory_path = f"table{query_counter}.csv"
            
            df_table.to_csv(saving_directory_path)
         
        # Update counter:
        self.query_counter = query_counter + 1    
            
        return df_table
    

    def create_new_view (self, view_id, query, show_table = True, export_csv = False, saving_directory_path = ""):
        """
        Creates a view in Google BigQuery if a view with the same name does not already exist.
        
        Important! This function uses a connection from google cloud sdk already configured

        Parameters:
        : param: view_id (str): The ID of the view to be created. If no ID is provided, a table is created
        : param: query (str): The SQL query defining the view.

        Returns:
        None
        """

        from google.cloud import bigquery

        view_id, query = str(view_id), str(query)
        # self.project = project
        bqclient, bqstorageclient = self.bqclient, self.bqstorageclient
        
        # Check if the view already exists
        if self.table_exists(bqclient, view_id):
            print(f"A view with the ID '{view_id}' already exists. Not creating a new view.")
            return self

        
        df = (
            bqclient.query(query)
            .result()
            .to_dataframe()
        )

        # This step creates the table for specified view
        view = bigquery.Table(view_id)
        view.view_query = query

        # Make an API request to create the view.
        view = bqclient.create_table(view)

        query_counter = self.query_counter
        
        if ControlVars.show_results: # dominant context
            if (show_table): 
                print("Returned view:\n")
                try:
                    from IPython.display import display
                    display(df)
                    
                except:
                    print(df)
        
        # Vars function allows accessing the attributes from a class as a key from a dictionary.
        # vars(object) is a dictionary where each key is one attribute from the object. A new attribute may be
        # created by setting a value for a new key;
        # Then, an attribute name may be created as a string:
        vars(self)[f"df_view{query_counter}"] = df

        if (export_csv):
            if ((saving_directory_path is None)|(saving_directory_path == '')):
                saving_directory_path = f"table{query_counter}.csv"
            
            df.to_csv(saving_directory_path)
         
        # Update counter:
        self.query_counter = query_counter + 1 

        return df


class IngestExcelTables:
    """
    Class for picking Excel files with non-structured data saved into several tables.
    For this class to work, the tables do not need to be input in a structured format, but they must be formatted
    as tables. With that, they can be detected and converted to Pandas dataframes. Alternatively, if no table is
    detected, the whole sheet is loaded

    : param: file_path (str): full path where the Excel file is locally stored. The file extension must be provided 
    (xlsx, etc).

    """
    
    def __init__ (self, file_path):
        
        self.file_path = file_path
        # Load workbook Openpyxl documentation:
        # https://openpyxl.readthedocs.io/en/stable/worksheet_tables.html?highlight=load_workbook#table-as-a-print-area
        self.wb = openpyxl.load_workbook(file_path)
        self.worksheets = self.wb.worksheets # list of worksheets
        self.loaded_dfs = [] # list of loaded dataframes
    
    
    def pre_cleansing(self, df):
        """Pre-cleansing of the dataframe. The methods are highly prone to result in duplicate rows and 
        completely blank rows or columns"""
        
        # Remove completely blank columns:
        df = df.dropna(axis = 1, how = 'all')
        # Remove completely blank rows:
        df = df.dropna(axis = 0, how = 'all')
        # Drop duplicates:
        df = df.drop_duplicates()
        # Reset index:
        df = df.reset_index(drop = True)
        
        return df
    
    
    def get_table_parameters(self, tab_range):
        
        """Use the table ranges identified by Openpyxl to obtain the
        parameters skiprows and usecols from pd.read_excel function.
        
        skiprows (int): indicates row much rows to skip.
        usecols (str): indicates the columns to use, in format "A:B"
        
        The tab_range obtained with openpyxl, in turns, is a string as "A10:BC78", which
        may contain an arbitrary number of letters, followed by an arbitrary number of digits.
        """
        
        # Split string as "A10:BC78" in ":". It will create a list with two strings ["A10", "BC78"]
        ranges = tab_range.split(":")
        
        # Iterate through characters in the first string from list ranges
        for i in range(len(ranges[0])):
            try:
                # If it was possible to break the string in two and convert the second part to
                # integer, thus the process may be finished:
                first_col, first_row = ranges[0][:i], int(ranges[0][i:])
                break
            
            except:
                # the error is raised if it was not possible to convert to int (e.g. 'A1') is not
                # conversible. So pass, and try again.
                pass
        
        # Repeat the process for the second string
        for i in range(len(ranges[1])):
            try:
                last_col, last_row = ranges[1][:i], int(ranges[1][i:])
                break
            
            except:
                pass
        
        # Concatenate strings to obtain usecols parameter:
        usecols = first_col + ":" + last_col
        
        # Parameter skipcols will be one unit less tha first_row
        # For instance, "A10" starts on the tenth row, so 9 rows must be skipped:
        skiprows = first_row - 1
        
        return usecols, skiprows

        
    def read_table(self, sheet_name, tab_range, has_header = True):
        
        # Get table parameters for pd.read_excel function:
        usecols, skiprows = self.get_table_parameters(tab_range)
        
        # Read table as pandas dataframe:
        
        if (has_header == True):         
            table = pd.read_excel(self.file_path, sheet_name = sheet_name, skiprows = skiprows, usecols = usecols, na_values = None, verbose = False, parse_dates = True)
        
        else:
            table = pd.read_excel(self.file_path, sheet_name = sheet_name, header = None, skiprows = skiprows, usecols = usecols, na_values = None, verbose = False, parse_dates = True)
        
        # Do the pre-cleansing:
        table = self.pre_cleansing(table)
            
        return table
    
    
    def read_full_sheet(self, sheet_name, has_header = True):
        """Read the entire sheet, instead of an individual table"""
        
        if (has_header == True):         
            table = pd.read_excel(self.file_path, sheet_name = sheet_name, na_values = None, verbose = False, parse_dates = True)

        else:
            table = pd.read_excel(self.file_path, sheet_name = sheet_name, header = None, na_values = None, verbose = False, parse_dates = True)
                    
        # Pre-cleansing:
        table = self.pre_cleansing(table)
        
        return table

    
    def load_dfs(self, has_header = True):
        
        loaded_dfs = self.loaded_dfs
        
        for ws in self.worksheets:
            sheet_name = ws.title
            
            # ws.tables is a dictionary containing the tables in a given sheet. If the dictionary is empty,
            # (len = 0), load the whole sheet as a dataframe:
            if (len(ws.tables) == 0):
                table = self.read_full_sheet(sheet_name, has_header)
                # Store on the list:
                loaded_dfs.append({'sheet': sheet_name, 'table': sheet_name, 'df': table})
                
            
            else:
                # Loop through each table:
                for table_name, tab_values in zip(ws.tables, ws.tables.values()):
                    tab_range = tab_values.ref

                    # Read the table:
                    table = self.read_table(sheet_name, tab_range, has_header)
                    # Store on the list:
                    loaded_dfs.append({'sheet': sheet_name, 'table': table_name, 'df': table})
        
        
        # Save list as attribute:
        self.loaded_dfs = loaded_dfs
                
        return self
    
    
    def export_processed_excel_file(self):
        
        file_path = "processed_excel.xlsx"
        
        try:
            # The replacement of a Sheet will only occur in the append ('a') mode.
            # 'a' is a mode available for the cases where an Excel file is already present.
            # Let's check if there is an Excel file previously created, so that we will not
            # delete it:
            with pd.ExcelWriter(file_path, date_format = "YYYY-MM-DD",
                                datetime_format = "YYYY-MM-DD HH:MM:SS",
                                mode = 'a', if_sheet_exists = 'replace') as writer:
                
                for loaded_df in self.loaded_dfs:
                    df, table = loaded_df['df'], loaded_df['table']
                    
                    if (table != loaded_df['sheet']):
                        table = loaded_df['sheet'] + "_" + table
                        
                    df.to_excel(writer, sheet_name = table, na_rep='', 
                                    header = True, index = False, 
                                    startrow = 0, startcol = 0, merge_cells = False, 
                                    inf_rep = 'inf')
        
        except:
            # The context manager created by class ExcelWriter with 'a' mode returns an error when
            # there is no Excel file available. Since we do not have the risk of overwriting the file,
            # we can open the writer in write ('w') mode to create a new spreadsheet:
            with pd.ExcelWriter(file_path, date_format = "YYYY-MM-DD",
                                datetime_format = "YYYY-MM-DD HH:MM:SS", mode = 'w') as writer:   
                
                for loaded_df in self.loaded_dfs:
                    df, table = loaded_df['df'], loaded_df['table']
                    
                    if (table != loaded_df['sheet']):
                        table = loaded_df['sheet'] + "_" + table
                        
                    df.to_excel(writer, sheet_name = table, index = False, 
                                startrow = 0, startcol = 0, merge_cells = False, 
                                inf_rep = 'inf')
        
        
        if ControlVars.show_results: 
            print(f"Dataframes exported as Excel file to notebook\'s workspace as \'{file_path}\'.")
            print("Warning: if there was a sheet with the same name as the exported ones, it was replaced by the exported dataframe.")
        
        
    def ingestion_pipeline(self, has_header = True, export_excel = True):
        
        self = self.load_dfs(has_header)
        
        if export_excel:
            self.export_processed_excel_file()
        
        return self


class SharePointDownloader:
    """Pipeline for accessing SharePoint files"""

    def __init__(self):

        import os
        from dotenv import load_dotenv
        from msal import ConfidentialClientApplication
        load_dotenv()

        # Environment variables loading
        self.CLIENT_ID = os.environ.get('CLIENT_ID')
        self.CLIENT_SECRET = os.environ.get('CLIENT_SECRET')
        self.TENANT_ID = os.environ.get('TENANT_ID')
        self.SHAREPOINT_SITE_ID = os.environ.get('SHAREPOINT_SITE_ID')
        
        # This url is used at each request, as it is the main request url
        self.url = f'https://graph.microsoft.com/v1.0/sites/{self.SHAREPOINT_SITE_ID}/drives/'
            
        # Authentication Config
        self.authority = 'https://login.microsoftonline.com/' + self.TENANT_ID
        self.scope = ['https://graph.microsoft.com/.default']

        # APP Initialization
        self.app = ConfidentialClientApplication(
            self.CLIENT_ID, 
            authority = self.authority,
            client_credential = self.CLIENT_SECRET
        )
    

    def get_token(self):
        """
        This function aims to create a 'headers' object to be used at each get request on the Graph API.
        The return is a json/dictionary object to be read by the headers argument, as it can be seen in the functions below.
        """
        from dotenv import load_dotenv
        load_dotenv()

        try:
            result = self.app.acquire_token_for_client(scopes = self.scope)
            access_token = result['access_token']
            headers = {'Authorization': 'Bearer ' + access_token}
            return headers
        
        except Exception as e:
            print(f'Error: {e}')
        
    
    def get_response_id(self, result_json, match):
        """
            This function is responsible to read all API response and return the folder/file IDs.
            It converts the response from get() method into a JSON format and gets the intended ID by matching the subsequent name on arg.
            
            The arguments of this function are:
            - result_json: inputs the obtject as type requests.models.Response, which is converted to json and read by the function
            - match: string of folder/file to be found by the function. 
            
            If the name is not match as required, the return is none.
        """

        import json
        from dotenv import load_dotenv
        load_dotenv()

        try:
            # Converting the response into JSON format
            result_json = json.loads(result_json.content)
            
            # Verifying the id for the object
            for item in result_json['value']:
                if item['name'] == match:
                    return item['id']
            return None
        
        except Exception as e:
            print(f'Error: {e}')
    

    def get_drive_id(self, main_sharepoint_directory, path):
        """
        : param: main_sharepoint_directory (str): represents the primary location for fetching the information.
            Usually, this information is present in a location called 'Documents'.
        
        : param: path (str): path within the main directory. Example: if you want to access a folder called "SPC data"
            in the 'Documents' directory, path = 'SPC data'. If you want to access the content from 'Documents', keep
            path = None
        """
        import requests
        from dotenv import load_dotenv

        load_dotenv()
        
        try: 
            headers = self.get_token()
            response = requests.get(self.url, headers = headers)
            drive_id = self.get_response_id(result_json = response, match = main_sharepoint_directory)
            
            if path: # run if it is not None
                drive_response = requests.get(self.url + f'{drive_id}/root/children', headers = headers)
                root_folder_id = self.get_response_id(result_json = drive_response, match = path)
                
            return drive_id, root_folder_id
        
        except Exception as e:
            print(f'Error: {e}')


    def find_file(self, drive_id: str, folder_id: str, target_file_name: str, headers):
        """
            find_file is a API scraping function to find the ID of the target file on sharepoint.
            
            The arguments of this function are:
            - drive_id: string, the ID of the root folder on the API.
            - folder_id: string, uses the IDs of the folders from the root folder to search the target file.
            - target_file_name: string, the name of file to be searched. If found, stops the search.
            - headers: JSON, the authorization to run the get() request method.
            
            If the name is not match as required, the return is none.
        """
        import requests
        from dotenv import load_dotenv

        load_dotenv()

        try:
            # Requests to search the file ID.
            folder_url = f"{self.url}/{drive_id}/items/{folder_id}/children"
            folder_response = requests.get(folder_url, headers = headers)
            folder_result = folder_response.json()
            
            # This loops looks for the Download URL. If do not finds the file in the loop, it looks inside each folder.
            for item in folder_result['value']:
                if '@microsoft.graph.downloadUrl' in item:  # Check if it is a valid file
                    if item['name'] == target_file_name:
                        return item['id']
                else:
                    if 'folder' in item:
                        sub_folder_id = item['id']
                        file_id = self.find_file(drive_id, sub_folder_id, target_file_name, headers)
                        if file_id:
                            return file_id
            return None
        
        except Exception as e:
            print(f'Error: {e}')


    def download_file(self, target_file_name, main_sharepoint_directory = 'Documents', path = None):
        """
            : param: main_sharepoint_directory (str): represents the primary location for fetching the information.
            Usually, this information is present in a location called 'Documents'.
        
            : param: path (str): path within the main directory. Example: if you want to access a folder called "SPC data"
                in the 'Documents' directory, path = 'SPC data'. If you want to access the content from 'Documents', keep
                path = None

            The function is able to read and download the sharepoint hosted file from the result given the file ID obtained by the method find_file()
            
            The arguments of this function are:
            - file_id: string, the ID of the root folder on the API.
            - target_file_name: string, the name of file to be searched. If found, stops the search.
            - headers: JSON, the authorization to run the get() request method.
            
            If the name is not match as required, the return is none.
        """
        import requests
        from urllib.request import urlretrieve
        from dotenv import load_dotenv

        load_dotenv()

        try:
            # Getting each important ID to download the file
            headers = self.get_token()
            drive_id, root_folder_id = self.get_drive_id(main_sharepoint_directory, path)
            file_id = self.find_file(drive_id, root_folder_id, target_file_name, headers)
            
            # Requests the Download URL from the API and downloads it within the system set up.
            if file_id:
                file_url = f"{self.url}/{drive_id}/items/{file_id}"
                file_result = requests.get(file_url, headers = headers).json()
                file_download_url = file_result["@microsoft.graph.downloadUrl"]
                urlretrieve(file_download_url, file_result['name'])
                print("Download successful!")
            else:
                raise InvalidInputsError("File not found.")
                
        except Exception as e:
            print(f'Error: {e}')
        # sharepoint_downloader = SharePointDownloader()
        # sharepoint_downloader.download_file(target_file_name = "data.xlsx")

